"""
Parser dei file di CARICO / RITIRO merce per magazzino.

Ogni magazzino ha il proprio formato file (Excel/PDF). Il modulo:
1. Auto-rileva quale magazzino è dal contenuto (header signature)
2. Parsa righe in formato standard: {ean, qta, sap_order_number, articolo, ...}

Aggiungere un nuovo magazzino = creare due funzioni `_detect_X` e `_parse_X`,
poi registrarle in REGISTRO_CARICO (o REGISTRO_RITIRO).

Nota sui warehouse_id: il parser non conosce gli id veri del DB; ritorna
solo il NOME del magazzino (es. "BO - LTF LOGISTICS"). Sarà il backend
a fare il lookup nella collection `magazzini` per ottenere l'id reale,
così se in futuro l'utente rinomina il magazzino non si rompe nulla.
"""
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime as _dt
from openpyxl import load_workbook


def _norm_header(s) -> str:
    """Normalizza un'intestazione per match robusto: lowercase + collassa spazi."""
    if s is None:
        return ""
    return " ".join(str(s).strip().lower().split())


# ════════════════════════════════════════════════════════════════════════════
# BO - LTF LOGISTICS
# ════════════════════════════════════════════════════════════════════════════
# Formato Excel:
# - Header in riga 3
# - Colonne: A=N ORDINE ADIDAS, B=DOC_FORN, C=N ORDINE COBA, D=ARTICOLO,
#            E=DESCRIZIONE, F=EAN, G=MIS, H=Somma di pz
# - Ultima riga: "Totale complessivo" (da skippare)

# Signature: insieme di header che devono essere TUTTI presenti nella stessa
# riga per riconoscere il formato. Sufficientemente specifico da non
# confondersi con altri magazzini.
_BO_LTF_SIG = {
    "n ordine adidas",
    "doc_forn",
    "n ordine coba",
    "ean",
    "somma di pz",
}


def _find_header_row(xlsx_path: str, signature: set, max_scan: int = 10) -> Optional[Tuple[int, Dict[str, int]]]:
    """Cerca nelle prime `max_scan` righe una che contenga tutta la signature.
    Ritorna (row_idx_1based, {header_normalized: col_idx_0based}) o None."""
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
            if not row:
                continue
            normalized = [_norm_header(c) for c in row]
            if signature.issubset(set(normalized)):
                col_map = {h: i for i, h in enumerate(normalized) if h}
                wb.close()
                return (ridx, col_map)
        wb.close()
    except Exception:
        pass
    return None


def _detect_bo_ltf(xlsx_path: str) -> bool:
    """BO - LTF LOGISTICS: header in qualsiasi riga 1-10.
    Tutti i file di carico per questo magazzino usano lo stesso formato Adidas
    ma a volte l'intestazione parte dalla riga 3, a volte dalla riga 4 (titoli
    sopra). Quindi la posizione non è discriminante."""
    return _find_header_row(xlsx_path, _BO_LTF_SIG, max_scan=10) is not None


def _parse_bo_ltf(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parsa file di ricezione BO - LTF LOGISTICS.
    Ritorna lista di righe normalizzate: {ean, qta, sap_order_number, articolo, descrizione, mis, doc_forn, n_ordine_coba}."""
    header = _find_header_row(xlsx_path, _BO_LTF_SIG)
    if not header:
        raise ValueError("Header BO - LTF LOGISTICS non trovato")
    header_row, col_map = header

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    righe = []
    try:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(c is not None for c in row):
                continue
            # Skip riga "Totale complessivo"
            v_a = row[0] if len(row) > 0 else None
            if isinstance(v_a, str) and "totale" in v_a.lower():
                continue

            def g(key: str):
                i = col_map.get(key)
                if i is None or i >= len(row):
                    return None
                return row[i]

            ean = g("ean")
            qta = g("somma di pz")
            if not ean or qta is None:
                continue
            try:
                qta_int = int(float(qta))
            except (ValueError, TypeError):
                continue
            if qta_int <= 0:
                continue

            righe.append({
                "ean": str(ean).strip(),
                "qta": qta_int,
                "sap_order_number": str(g("n ordine adidas") or "").strip(),
                "articolo": str(g("articolo") or "").strip(),
                "descrizione": str(g("descrizione") or "").strip(),
                "mis": str(g("mis") or "").strip(),
                "doc_forn": str(g("doc_forn") or "").strip(),
                "n_ordine_coba": str(g("n ordine coba") or "").strip(),
            })
    finally:
        wb.close()
    return righe


# ════════════════════════════════════════════════════════════════════════════
# GIACENZA PLURI (multi-cliente, non-COBA)
# ════════════════════════════════════════════════════════════════════════════
# Formato Adidas: 1 foglio, header in riga 1.
# Colonne chiave: Sales document, EAN, Delivery quantity, CLIENTE.
# Logica: escludi righe con CLIENTE che contiene "COBA".
_PLURI_SIG = {
    "sales document",
    "ean",
    "delivery quantity",
    "cliente",
}

PLURI_WAREHOUSE = "__PLURI__"


def _detect_pluri(xlsx_path: str) -> bool:
    return _find_header_row(xlsx_path, _PLURI_SIG, max_scan=5) is not None


def _parse_pluri(xlsx_path: str) -> List[Dict[str, Any]]:
    """Ritorna solo le righe con DATA ARRIVO valorizzata (merce effettivamente arrivata)
    e CLIENTE != COBA. Ogni riga rappresenta un cartone/lotto; l'aggregazione per
    EAN+SAP avviene nel backend prima del match con gli ordini."""
    header = _find_header_row(xlsx_path, _PLURI_SIG)
    if not header:
        raise ValueError("Header giacenza pluri non trovato")
    header_row, col_map = header

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    righe = []
    try:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(c is not None for c in row):
                continue

            def g(key: str):
                i = col_map.get(key)
                if i is None or i >= len(row):
                    return None
                return row[i]

            # Solo righe con DATA ARRIVO = datetime reale.
            # Celle con formule che ritornano '#N/A', '' o None vengono scartate.
            data_arrivo = g("data arrivo")
            if not isinstance(data_arrivo, _dt):
                continue

            cliente_raw = g("cliente")
            if cliente_raw is not None and "COBA" in str(cliente_raw).strip().upper():
                continue

            ean = g("ean")
            qta = g("delivery quantity")
            if ean is None or qta is None:
                continue
            try:
                qta_int = int(float(qta))
            except (ValueError, TypeError):
                continue
            if qta_int <= 0:
                continue

            ean_str = str(ean).strip()
            if ean_str.endswith(".0"):
                ean_str = ean_str[:-2]

            sap = g("sales document")
            sap_str = str(sap).strip() if sap is not None else ""
            if sap_str.endswith(".0"):
                sap_str = sap_str[:-2]

            righe.append({
                "ean": ean_str,
                "qta": qta_int,
                "sap_order_number": sap_str,
                "articolo": str(g("material") or "").strip(),
                "descrizione": str(g("material description") or "").strip(),
                "mis": str(g("grid value") or "").strip(),
                "cliente_pluri": str(cliente_raw or "").strip(),
            })
    finally:
        wb.close()
    return righe


# ════════════════════════════════════════════════════════════════════════════
# REGISTRO PARSER CARICO (ingresso merce)
# ════════════════════════════════════════════════════════════════════════════
# Lista di (warehouse_name, detect_fn, parse_fn).
# Aggiungere qui i nuovi magazzini.
REGISTRO_CARICO: List[Tuple[str, Any, Any]] = [
    ("BO - LTF LOGISTICS", _detect_bo_ltf, _parse_bo_ltf),
    (PLURI_WAREHOUSE, _detect_pluri, _parse_pluri),
]


def detect_carico(xlsx_path: str) -> Optional[Tuple[str, List[Dict[str, Any]]]]:
    """Auto-rileva quale magazzino corrisponde al file e parsa le righe.
    Ritorna (warehouse_name, righe) o None se nessun parser riconosce il file."""
    for name, detect_fn, parse_fn in REGISTRO_CARICO:
        if detect_fn(xlsx_path):
            return (name, parse_fn(xlsx_path))
    return None


def parse_carico_for(xlsx_path: str, warehouse_name: str) -> List[Dict[str, Any]]:
    """Forza il parser di un magazzino specifico (override dell'auto-detect)."""
    for name, _, parse_fn in REGISTRO_CARICO:
        if name == warehouse_name:
            return parse_fn(xlsx_path)
    raise ValueError(f"Nessun parser carico per magazzino '{warehouse_name}'")


# ════════════════════════════════════════════════════════════════════════════
# BO - LTF LOGISTICS — RITIRO / SPEDIZIONE (uscita merce)
# ════════════════════════════════════════════════════════════════════════════
# Formato file (es. "Taranto, 13-04-26.xlsx"):
# - Sheet "dettaglio dn", header in riga 3
# - Colonne: A=Partenza, B=N° ordine (interno, NON SAP), C=Delivery,
#            D=Material (articolo), E=Grid Value2 (size),
#            F=Somma di Delivery quantity
# - CONVENZIONE: l'utente colora di giallo (FFFFFF00) la cella della colonna D
#   delle righe che effettivamente vengono spedite/ritirate.
#   Le righe NON gialle vanno IGNORATE.

_BO_LTF_RITIRO_SHEET = "dettaglio dn"
_BO_LTF_RITIRO_SIG = {
    "partenza",
    "n° ordine",
    "delivery",
    "material",
    "grid value2",
    "somma di delivery quantity",
}
_YELLOW_FILLS = {"FFFFFF00", "FFFFFFFF00"}  # giallo con alpha


def _is_yellow(cell) -> bool:
    """True se la cella ha fill giallo. Robusto: alcune varianti di Excel
    salvano l'alpha in modo diverso, accettiamo entrambe."""
    try:
        fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    except Exception:
        return False
    if not fg:
        return False
    fg_str = str(fg).upper()
    # Match: deve finire in "FFFF00" e iniziare con FF (alpha) o senza
    return fg_str.endswith("FFFF00") and ("FFFF00" in fg_str)


def _detect_bo_ltf_ritiro(xlsx_path: str) -> bool:
    """True se il file contiene lo sheet 'dettaglio dn' con la signature attesa."""
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        if _BO_LTF_RITIRO_SHEET not in wb.sheetnames:
            wb.close()
            return False
        ws = wb[_BO_LTF_RITIRO_SHEET]
        # Scansiono prime 6 righe per trovare header
        found = False
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            if not row:
                continue
            headers = {_norm_header(c) for c in row if c is not None}
            if _BO_LTF_RITIRO_SIG.issubset(headers):
                found = True
                break
        wb.close()
        return found
    except Exception:
        return False


def _parse_bo_ltf_ritiro(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parsa lo sheet 'dettaglio dn' del file di ritiro BO - LTF LOGISTICS.
    Ritorna SOLO le righe in cui la cella della colonna 'Material' è gialla.

    Ogni riga: {articolo, size, qta, n_ordine_interno, delivery, partenza}.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if _BO_LTF_RITIRO_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{_BO_LTF_RITIRO_SHEET}' non trovato")
    ws = wb[_BO_LTF_RITIRO_SHEET]

    # Trovo header row e mappa colonne
    header_row_idx = None
    col_map = {}
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=6), 1):
        normalized = [_norm_header(c.value) for c in row]
        if _BO_LTF_RITIRO_SIG.issubset(set(normalized)):
            header_row_idx = ridx
            for i, h in enumerate(normalized):
                col_map[h] = i
            break
    if header_row_idx is None:
        wb.close()
        raise ValueError("Header ritiro BO - LTF LOGISTICS non trovato")

    material_col = col_map["material"]
    righe = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        if not row or material_col >= len(row):
            continue
        cd = row[material_col]
        # Solo righe con Material giallo
        if cd.value is None or not _is_yellow(cd):
            continue

        def g(key: str):
            i = col_map.get(key)
            if i is None or i >= len(row):
                return None
            return row[i].value

        articolo = cd.value
        size = g("grid value2")
        qta = g("somma di delivery quantity")
        if not articolo or qta is None:
            continue
        try:
            qta_int = int(float(qta))
        except (ValueError, TypeError):
            continue
        if qta_int <= 0:
            continue

        # EAN: colonna opzionale. Se presente e valorizzata, è il modo più
        # affidabile di matchare l'articolo (univoco). Se vuota, fallback
        # (articolo + size) lato server. Accetto sia "ean" che "EAN".
        ean_val = g("ean")
        ean_str = ""
        if ean_val is not None:
            ean_str = str(ean_val).strip()
            # Excel a volte legge un numero lungo come float "4068816832815.0"
            # → ripristino la forma intera senza decimali.
            if ean_str.endswith(".0"):
                ean_str = ean_str[:-2]

        righe.append({
            "articolo": str(articolo).strip(),
            "size": str(size).strip() if size is not None else "",
            "qta": qta_int,
            "ean": ean_str,  # "" se la cella non c'è o è vuota
            "n_ordine_interno": str(g("n° ordine") or "").strip(),
            "delivery": str(g("delivery") or "").strip(),
            "partenza": str(g("partenza") or "").strip(),
        })
    wb.close()
    return righe


# ════════════════════════════════════════════════════════════════════════════
# DDT NEXT (ritiro cliente da file DDT)
# ════════════════════════════════════════════════════════════════════════════
# Formato: foglio "DETTAGLIO", header in riga 1.
# Colonne chiave: EAN (I), Delivery quantity (H), Purchase order no. (C),
#                 Data arrivo (N), Rif ns ddt (Q), Material (E), Grid Value (F).
_DDT_NEXT_SIG = {"ean", "delivery quantity", "rif ns ddt"}
DDT_NEXT_WAREHOUSE = "__DDT_NEXT__"
_DDT_NEXT_SHEET = "DETTAGLIO"


def _detect_ddt_next(xlsx_path: str) -> bool:
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        if _DDT_NEXT_SHEET not in wb.sheetnames:
            wb.close()
            return False
        ws = wb[_DDT_NEXT_SHEET]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if not row:
                continue
            headers = {_norm_header(c) for c in row if c is not None}
            if _DDT_NEXT_SIG.issubset(headers):
                found = True
                break
        wb.close()
        return found
    except Exception:
        return False


def _parse_ddt_next(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parsa foglio DETTAGLIO del DDT NEXT.
    Ogni riga: {ean, qta, sap_order_number, data_ritiro, ddt_ref, articolo, mis, descrizione}."""
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[_DDT_NEXT_SHEET]
    except Exception as e:
        raise ValueError(f"Impossibile aprire foglio {_DDT_NEXT_SHEET}: {e}")

    header_row_idx = None
    col_map: Dict[str, int] = {}
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if not row:
            continue
        normalized = [_norm_header(c) for c in row]
        if _DDT_NEXT_SIG.issubset(set(normalized)):
            header_row_idx = ridx
            col_map = {h: i for i, h in enumerate(normalized) if h}
            break
    if header_row_idx is None:
        wb.close()
        raise ValueError("Header DDT NEXT non trovato nel foglio DETTAGLIO")

    righe = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not any(c is not None for c in row):
            continue

        def g(key: str):
            i = col_map.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        ean = g("ean")
        qta = g("delivery quantity")
        if ean is None or qta is None:
            continue
        try:
            qta_int = int(float(qta))
        except (ValueError, TypeError):
            continue
        if qta_int <= 0:
            continue

        ean_str = str(ean).strip()
        if ean_str.endswith(".0"):
            ean_str = ean_str[:-2]

        sap_raw = g("purchase order no.")
        sap_str = str(sap_raw).strip() if sap_raw is not None else ""
        if sap_str.endswith(".0"):
            sap_str = sap_str[:-2]

        # DATA PAR = data partenza/ritiro (colonna P).
        # DATA ARRIVO (colonna N) è la data di arrivo in magazzino — non usarla.
        data_raw = g("data par")
        if isinstance(data_raw, _dt):
            data_str = data_raw.strftime("%Y-%m-%d")
        elif data_raw:
            data_str = str(data_raw).strip()
        else:
            data_str = ""

        ddt_ref = str(g("rif ns ddt") or "").strip()

        righe.append({
            "ean": ean_str,
            "qta": qta_int,
            "sap_order_number": sap_str,
            "data_ritiro": data_str,
            "ddt_ref": ddt_ref,
            "articolo": str(g("material") or "").strip(),
            "mis": str(g("grid value") or "").strip(),
            "descrizione": str(g("material description") or "").strip(),
        })
    wb.close()
    return righe


# ════════════════════════════════════════════════════════════════════════════
# REGISTRO PARSER RITIRO/SPEDIZIONE (uscita merce)
# ════════════════════════════════════════════════════════════════════════════
REGISTRO_RITIRO: List[Tuple[str, Any, Any]] = [
    ("BO - LTF LOGISTICS", _detect_bo_ltf_ritiro, _parse_bo_ltf_ritiro),
    (DDT_NEXT_WAREHOUSE, _detect_ddt_next, _parse_ddt_next),
]


def detect_ritiro(xlsx_path: str) -> Optional[Tuple[str, List[Dict[str, Any]]]]:
    """Auto-rileva il magazzino del file di ritiro/spedizione + parsa le righe gialle."""
    for name, detect_fn, parse_fn in REGISTRO_RITIRO:
        if detect_fn(xlsx_path):
            return (name, parse_fn(xlsx_path))
    return None
