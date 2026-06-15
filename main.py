"""
Backend FastAPI per Logistic-APP — gestione logistica B2B abbigliamento.

Modello dati:
- Utente: admin (te) o customer (cliente B2B, agganciato a sold_to_code)
- Cliente: sold_to (es. "Esposito")
- Magazzino: name + street + city, con zone (codici Ship To)
- OrdineCliente: SAP order number + personal reference (nome) + righe
- Riga: EAN univoco (taglia) + Article ID (modello+colore) + qty
- Carico magazzino: evento che incrementa arrived_qty di una riga
- Ritiro: evento che incrementa picked_up_qty di una riga

Stati derivati:
- Arrivo riga:  in_attesa | parziale | completo (su arrived_qty vs expected_qty)
- Ritiro riga:  non_ritirato | parziale | completo (su picked_up_qty vs expected_qty)

Storage:
- Default: JSON file su disco (data/store.json) — zero setup
- Produzione: MongoDB se MONGODB_URI è settato come env var

Stack: FastAPI + uvicorn + (motor opzionale) + bcrypt + PyJWT
"""

import os
import re
import json
import uuid
import asyncio
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional, List, Dict, Any
from contextlib import asynccontextmanager

import bcrypt
import jwt
from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr, Field

# ── CONFIG ──────────────────────────────────────────────────────────────
ROOT_DIR = Path(__file__).parent
DATA_DIR = ROOT_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
STORE_FILE = DATA_DIR / "store.json"
SEED_FILE = DATA_DIR / "seed.json"

JWT_SECRET = os.environ.get("JWT_SECRET", "dev-secret-change-in-production-please")
MONGODB_URI = os.environ.get("MONGODB_URI", "")  # se vuoto → storage JSON
CORS_ORIGINS = [o.strip() for o in os.environ.get("CORS_ORIGINS", "*").split(",")]

ACCESS_TTL = timedelta(hours=24)
REFRESH_TTL = timedelta(days=7)
REMEMBER_TTL = timedelta(days=30)


# ── STORAGE LAYER (JSON fallback / Mongo) ─────────────────────────────────
# Astrazione minima: ogni "collezione" è una lista di dict. Le operazioni base
# sono find / insert / update / delete. In JSON mode tutto vive in memoria
# e si serializza su file ad ogni write. In Mongo mode delega al driver motor.

class JSONStore:
    """Storage backed da un singolo file JSON. Adatto per MVP, demo, single-node."""

    def __init__(self, path: Path):
        self.path = path
        self._lock = asyncio.Lock()
        self._data = self._load()

    def _load(self) -> dict:
        if self.path.exists():
            try:
                return json.loads(self.path.read_text(encoding="utf-8"))
            except Exception as e:
                print(f"[storage] WARN: store.json corrotto, ricomincio da capo: {e}")
        return {
            "users": [],
            "clienti": [],
            "magazzini": [],
            "ordini": [],
            "carichi": [],
            "ritiri": [],
            "tickets": [],
        }

    async def _save(self):
        # Scrittura atomica: tmp → rename
        tmp = self.path.with_suffix(".tmp")
        tmp.write_text(json.dumps(self._data, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(self.path)

    async def find(self, col: str, query: Optional[dict] = None) -> List[dict]:
        """Filtro by-equality (semplice): query={"sold_to": "Esposito"}"""
        items = self._data.get(col, [])
        if not query:
            return list(items)
        return [x for x in items if all(x.get(k) == v for k, v in query.items())]

    async def find_one(self, col: str, query: dict) -> Optional[dict]:
        for x in self._data.get(col, []):
            if all(x.get(k) == v for k, v in query.items()):
                return x
        return None

    async def insert(self, col: str, doc: dict) -> dict:
        async with self._lock:
            self._data.setdefault(col, []).append(doc)
            await self._save()
        return doc

    async def update_one(self, col: str, query: dict, patch: dict) -> Optional[dict]:
        async with self._lock:
            for x in self._data.get(col, []):
                if all(x.get(k) == v for k, v in query.items()):
                    x.update(patch)
                    await self._save()
                    return x
        return None

    async def delete_one(self, col: str, query: dict) -> bool:
        async with self._lock:
            items = self._data.get(col, [])
            for i, x in enumerate(items):
                if all(x.get(k) == v for k, v in query.items()):
                    items.pop(i)
                    await self._save()
                    return True
        return False

    async def delete_many(self, col: str, query: Optional[dict] = None) -> int:
        """Cancella tutti i doc che matchano query (o tutti se query è None/empty).
        Ritorna il numero di doc cancellati."""
        async with self._lock:
            items = self._data.get(col, [])
            if not query:
                n = len(items)
                self._data[col] = []
                await self._save()
                return n
            kept = [x for x in items if not all(x.get(k) == v for k, v in query.items())]
            n_deleted = len(items) - len(kept)
            if n_deleted:
                self._data[col] = kept
                await self._save()
            return n_deleted

    async def bulk_upsert(self, col: str, docs: list, key_field: str) -> dict:
        """Versione in-memory del bulk upsert (per JSONStore fallback)."""
        upserted = 0
        modified = 0
        async with self._lock:
            items = self._data.setdefault(col, [])
            by_key = {x.get(key_field): i for i, x in enumerate(items) if x.get(key_field) is not None}
            for d in docs:
                k = d.get(key_field)
                if k in by_key:
                    items[by_key[k]] = dict(d)
                    modified += 1
                else:
                    items.append(dict(d))
                    by_key[k] = len(items) - 1
                    upserted += 1
            await self._save()
        return {"upserted": upserted, "modified": modified}

    async def next_serial(self, name: str) -> int:
        """Incrementa atomicamente un contatore (es. "tickets") e ritorna il
        nuovo valore. Garantito monotono anche con richieste concorrenti."""
        async with self._lock:
            counters = self._data.setdefault("_counters", {})
            val = counters.get(name, 0) + 1
            counters[name] = val
            await self._save()
        return val

    async def seed_if_empty(self):
        """Se non ci sono ordini E esiste un seed.json, lo importa."""
        if self._data.get("ordini"):
            return
        if not SEED_FILE.exists():
            return
        seed = json.loads(SEED_FILE.read_text(encoding="utf-8"))
        async with self._lock:
            self._data["clienti"] = seed.get("clienti", [])
            self._data["magazzini"] = seed.get("magazzini", [])
            self._data["ordini"] = seed.get("ordini", [])
            await self._save()
        print(f"[storage] Seed importato: {len(self._data['ordini'])} ordini")


class MongoStore:
    """Storage backed da MongoDB. Stessa interfaccia di JSONStore così il resto
    del codice non se ne accorge. Indicato per produzione / multi-utente."""

    def __init__(self, uri: str):
        # Import qui per non rompere il boot se motor non è installato
        from motor.motor_asyncio import AsyncIOMotorClient
        import certifi
        # tlsCAFile=certifi.where() è necessario per MongoDB Atlas su molte
        # piattaforme Linux (incluse Render) dove i CA di sistema non bastano
        self._client = AsyncIOMotorClient(uri, tlsCAFile=certifi.where())
        self._db = self._client.get_default_database()
        # Per coerenza con JSONStore: niente _id Mongo nelle risposte,
        # usiamo sempre il campo "id" / "sap_order_number" / "sold_to" come PK

    async def _strip_oid(self, doc: dict) -> dict:
        """Rimuove _id ObjectId che FastAPI non sa serializzare."""
        if doc and "_id" in doc:
            d = dict(doc)
            d.pop("_id", None)
            return d
        return doc

    async def find(self, col: str, query: Optional[dict] = None) -> List[dict]:
        cursor = self._db[col].find(query or {})
        docs = await cursor.to_list(length=None)
        return [await self._strip_oid(d) for d in docs]

    async def find_one(self, col: str, query: dict) -> Optional[dict]:
        d = await self._db[col].find_one(query)
        return await self._strip_oid(d) if d else None

    async def insert(self, col: str, doc: dict) -> dict:
        await self._db[col].insert_one(dict(doc))  # copia per non scrivere _id sul doc originale
        return doc

    async def update_one(self, col: str, query: dict, patch: dict) -> Optional[dict]:
        await self._db[col].update_one(query, {"$set": patch})
        return await self.find_one(col, query)

    async def delete_one(self, col: str, query: dict) -> bool:
        res = await self._db[col].delete_one(query)
        return res.deleted_count > 0

    async def delete_many(self, col: str, query: Optional[dict] = None) -> int:
        """Cancella tutti i doc che matchano query (o tutti se query è vuota).
        Wrapper su Mongo delete_many."""
        res = await self._db[col].delete_many(query or {})
        return res.deleted_count

    async def bulk_upsert(self, col: str, docs: list, key_field: str) -> dict:
        """Esegue upsert in batch: per ogni doc, se esiste già un documento
        con quel valore di key_field lo sostituisce, altrimenti lo inserisce.
        Ritorna {"upserted": N, "modified": M}. ~10x più veloce di chiamate seriali.
        """
        if not docs:
            return {"upserted": 0, "modified": 0}
        from pymongo import ReplaceOne
        ops = [ReplaceOne({key_field: d[key_field]}, dict(d), upsert=True) for d in docs]
        res = await self._db[col].bulk_write(ops, ordered=False)
        return {
            "upserted": res.upserted_count,
            "modified": res.modified_count,
        }

    async def next_serial(self, name: str) -> int:
        """Incrementa atomicamente un contatore MongoDB ($inc) e ritorna il
        nuovo valore. Usa find_one_and_update con upsert per garantire atomicità."""
        from pymongo import ReturnDocument
        res = await self._db["_counters"].find_one_and_update(
            {"id": name},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=ReturnDocument.AFTER,
        )
        if res is None:
            # Fallback di sicurezza: rileggi il documento appena creato
            res = await self._db["_counters"].find_one({"id": name})
        return int(res["seq"])

    async def seed_if_empty(self):
        """Se non ci sono ordini E esiste un seed.json, lo importa.
        Su Mongo questo è importante solo al PRIMO deploy: dopo, i dati
        persistono e seed_if_empty diventa no-op.

        IMPORTANTE: seedo solo le collezioni effettivamente vuote.
        Se in passato avevo già seedato magazzini/clienti e poi ho
        cancellato solo gli ordini, NON devo reinserire magazzini/clienti
        (altrimenti finisco con doc duplicati a parità di id)."""
        count_ord = await self._db["ordini"].count_documents({})
        if count_ord > 0:
            return
        if not SEED_FILE.exists():
            return
        seed = json.loads(SEED_FILE.read_text(encoding="utf-8"))
        count_cli = await self._db["clienti"].count_documents({})
        count_mag = await self._db["magazzini"].count_documents({})
        if seed.get("clienti") and count_cli == 0:
            await self._db["clienti"].insert_many([dict(x) for x in seed["clienti"]])
        if seed.get("magazzini") and count_mag == 0:
            await self._db["magazzini"].insert_many([dict(x) for x in seed["magazzini"]])
        if seed.get("ordini"):
            await self._db["ordini"].insert_many([dict(x) for x in seed["ordini"]])
        print(f"[storage] Mongo seed importato: {len(seed.get('ordini',[]))} ordini")


# Istanza globale storage: Mongo se MONGODB_URI è settato, altrimenti JSON file.
# In locale per dev → MONGODB_URI vuoto → JSON. In produzione (Render) → Mongo.
if MONGODB_URI:
    print(f"[storage] Modalità MongoDB (host nascosto)")
    store = MongoStore(MONGODB_URI)
else:
    print(f"[storage] Modalità JSON file: {STORE_FILE}")
    store = JSONStore(STORE_FILE)


# ── HELPERS: password e JWT ──────────────────────────────────────────────
def _hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt(12)).decode()


def _verify_pw(pw: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), h.encode())
    except Exception:
        return False


def _make_token(user_id: str, ttl: timedelta, kind: str) -> str:
    payload = {
        "sub": user_id, "kind": kind,
        "iat": datetime.now(timezone.utc),
        "exp": datetime.now(timezone.utc) + ttl,
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def _decode_token(token: str) -> Optional[dict]:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except jwt.PyJWTError:
        return None


def _public_user(u: dict) -> dict:
    """Espone solo i campi pubblici. MAI password_hash."""
    return {
        "id": u["id"],
        "username": u.get("username", ""),
        "email": u.get("email", ""),
        "role": u.get("role", "customer"),
        "sold_to_code": u.get("sold_to_code"),
        "enabled": u.get("enabled", False),
        "created_at": u.get("created_at"),
    }


# ── AUTH DEPENDENCY ───────────────────────────────────────────────────────
async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Token mancante")
    token = authorization.split(" ", 1)[1]
    payload = _decode_token(token)
    if not payload or payload.get("kind") != "access":
        raise HTTPException(401, "Token non valido o scaduto")
    user = await store.find_one("users", {"id": payload["sub"]})
    if not user:
        raise HTTPException(401, "Utente non trovato")
    if not user.get("enabled", False):
        raise HTTPException(403, "Account non abilitato")
    return user


async def get_admin_user(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Solo amministratori")
    return user


# ── STATI DERIVATI ────────────────────────────────────────────────────────
def _arrival_state(arrived: int, expected: int) -> str:
    if expected <= 0 or arrived <= 0:
        return "wait"
    if arrived >= expected:
        return "done"
    return "partial"


def _pickup_state(picked_up: int, expected: int) -> str:
    if expected <= 0 or picked_up <= 0:
        return "wait"
    if picked_up >= expected:
        return "done"
    return "partial"


def _aggregate_state(states: List[str]) -> str:
    """Aggrega stati riga in stato ordine:
    - tutte wait → wait
    - tutte done → done
    - altrimenti → partial"""
    if not states:
        return "wait"
    if all(s == "wait" for s in states):
        return "wait"
    if all(s == "done" for s in states):
        return "done"
    return "partial"


def _enrich_order(order: dict) -> dict:
    """Aggiunge agli ordini: expected/arrival/pickup state per riga e aggregati.

    Convenzione expected_qty:
    - Se la riga ha "pezzi_confermati" (formato master Adidas), usa quello
      come "atteso" — riflette i pezzi effettivamente confermati dal fornitore
      dopo eventuali rifiuti/cancellazioni.
    - Altrimenti fallback su total_qty − rejected_qty (formato Ordine Prova).
    """
    out = dict(order)
    arrival_states = []
    pickup_states = []
    total_expected = 0
    total_arrived = 0
    total_picked_up = 0
    for r in out.get("righe", []):
        if "pezzi_confermati" in r:
            expected = max(r.get("pezzi_confermati", 0), 0)
        else:
            expected = max(r.get("total_qty", 0) - r.get("rejected_qty", 0), 0)
        arrived = r.get("arrived_qty", 0)
        picked_up = r.get("picked_up_qty", 0)
        r["expected_qty"] = expected
        r["arrival_state"] = _arrival_state(arrived, expected)
        r["pickup_state"] = _pickup_state(picked_up, expected)
        arrival_states.append(r["arrival_state"])
        pickup_states.append(r["pickup_state"])
        total_expected += expected
        total_arrived += arrived
        total_picked_up += picked_up
    out["arrival_state"] = _aggregate_state(arrival_states)
    out["pickup_state"] = _aggregate_state(pickup_states)
    out["total_expected"] = total_expected
    out["total_arrived"] = total_arrived
    out["total_picked_up"] = total_picked_up
    return out


# ── MODELS Pydantic ───────────────────────────────────────────────────────
class RegisterReq(BaseModel):
    username: str = Field(min_length=3, max_length=32)
    email: EmailStr
    password: str = Field(min_length=8, max_length=128)


class LoginReq(BaseModel):
    username_or_email: str
    password: str
    remember: bool = False


class UpdateUserReq(BaseModel):
    enabled: Optional[bool] = None
    role: Optional[str] = Field(None, pattern="^(admin|customer)$")
    sold_to_code: Optional[str] = None


class CaricoRiga(BaseModel):
    ean: str
    qta: int = Field(gt=0)


class CaricoReq(BaseModel):
    # order_id è la PK lato app (formato "sap::nome"). Lo stesso SAP nel master
    # può corrispondere a più ordini distinti, quindi non basta più il SAP.
    order_id: str
    magazzino_id: str
    data_carico: Optional[str] = None  # ISO date, default = oggi
    note: Optional[str] = None
    righe: List[CaricoRiga]


class RitiroReq(BaseModel):
    order_id: str
    magazzino_id: str
    data_ritiro: Optional[str] = None
    note: Optional[str] = None
    righe: List[CaricoRiga]


# ── APP LIFECYCLE ─────────────────────────────────────────────────────────
async def _migrate_ticket_serials() -> dict:
    """Backfill: assegna serial a tutti i ticket che ne sono privi.
    - Ordine: cronologico per created_at (ticket più vecchi → numeri più bassi).
    - Idempotente: se tutti i ticket hanno già il serial esce subito.
    - Non usa next_serial() per evitare dipendenze dalla collection _counters:
      calcola il punto di partenza dal max serial già presente, poi allinea
      il contatore al termine in modo sicuro.
    Ritorna un dict con il riepilogo dell'operazione."""
    tickets = await store.find("tickets")
    without = [t for t in tickets if not t.get("serial")]
    if not without:
        return {"updated": 0, "already_ok": len(tickets)}

    # Punto di partenza: max serial già assegnato (0 se nessuno)
    max_existing = 0
    for t in tickets:
        try:
            v = int(t.get("serial") or 0)
            if v > max_existing:
                max_existing = v
        except (ValueError, TypeError):
            pass

    without.sort(key=lambda t: t.get("created_at", ""))
    next_num = max_existing + 1
    for t in without:
        serial = str(next_num).zfill(5)
        await store.update_one("tickets", {"id": t["id"]}, {"serial": serial})
        next_num += 1

    # Allinea il contatore al nuovo massimo così i prossimi ticket
    # continuano in sequenza senza conflitti
    final_max = next_num - 1
    try:
        # Porta il contatore al valore corretto (sovrascrivendo quello attuale
        # se è inferiore, senza usare next_serial che incrementerebbe)
        await store.update_one("_counters", {"id": "tickets"}, {"seq": final_max})
    except Exception:
        # Se _counters non esiste ancora è ok: next_serial lo creerà al primo
        # ticket nuovo e partirà dal valore sbagliato, ma next_serial usa
        # $inc che partirà da 0. Inseriamo il documento se assente.
        try:
            existing = await store.find_one("_counters", {"id": "tickets"})
            if not existing:
                await store.insert("_counters", {"id": "tickets", "seq": final_max})
        except Exception:
            pass

    print(f"[migration] Seriali assegnati a {len(without)} ticket (da 00001 a {str(final_max).zfill(5)})")
    return {"updated": len(without), "already_ok": len(tickets) - len(without)}


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Seed iniziale da seed.json (se store vuoto)
    await store.seed_if_empty()
    # Backfill seriali ticket mancanti (migration idempotente)
    try:
        await _migrate_ticket_serials()
    except Exception as e:
        print(f"[migration] WARN: backfill seriali fallito: {e}")
    # Crea admin di default se non esiste nessun user
    if not await store.find("users"):
        admin_pw = os.environ.get("ADMIN_DEFAULT_PW", "admin1234")
        await store.insert("users", {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "email": "admin@logistic-app.local",
            "password_hash": _hash_pw(admin_pw),
            "role": "admin",
            "sold_to_code": None,
            "enabled": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        print(f"[startup] Admin di default creato: username=admin, password={admin_pw}")
        # Crea anche un customer di esempio per "Esposito"
        clienti = await store.find("clienti")
        if clienti:
            sold_to = clienti[0]["sold_to"]
            await store.insert("users", {
                "id": str(uuid.uuid4()),
                "username": sold_to.lower(),
                "email": f"{sold_to.lower()}@cliente.local",
                "password_hash": _hash_pw("cliente1234"),
                "role": "customer",
                "sold_to_code": sold_to,
                "enabled": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            print(f"[startup] Customer di esempio creato: username={sold_to.lower()}, password=cliente1234")
    yield


app = FastAPI(title="Logistic-APP Backend", version="1.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── HEALTH ────────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return {"ok": True, "service": "Logistic-APP", "version": "1.0"}


@app.get("/health")
async def health():
    users = await store.find("users")
    orders = await store.find("ordini")
    return {
        "ok": True,
        "users_count": len(users),
        "orders_count": len(orders),
        "storage": "json",
    }


# ── AUTH ──────────────────────────────────────────────────────────────────
@app.post("/auth/register")
async def register(req: RegisterReq):
    """Registrazione self-service: crea un customer DISABILITATO.
    L'admin deve abilitarlo e collegarlo a un sold_to_code prima dell'accesso."""
    if await store.find_one("users", {"username": req.username}):
        raise HTTPException(400, "Username già usato")
    if await store.find_one("users", {"email": req.email}):
        raise HTTPException(400, "Email già usata")
    user = {
        "id": str(uuid.uuid4()),
        "username": req.username,
        "email": req.email,
        "password_hash": _hash_pw(req.password),
        "role": "customer",
        "sold_to_code": None,
        "enabled": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await store.insert("users", user)
    return {"ok": True, "message": "Registrato. Attendi che l'admin abiliti l'account."}


@app.post("/auth/login")
async def login(req: LoginReq):
    user = await store.find_one("users", {"username": req.username_or_email}) \
        or await store.find_one("users", {"email": req.username_or_email})
    if not user or not _verify_pw(req.password, user.get("password_hash", "")):
        raise HTTPException(401, "Credenziali non valide")
    if not user.get("enabled", False):
        raise HTTPException(403, "Account non ancora abilitato dall'admin")
    ttl = REMEMBER_TTL if req.remember else ACCESS_TTL
    return {
        "access_token": _make_token(user["id"], ttl, "access"),
        "refresh_token": _make_token(user["id"], REFRESH_TTL if not req.remember else REMEMBER_TTL, "refresh"),
        "user": _public_user(user),
    }


@app.post("/auth/refresh")
async def refresh(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Token mancante")
    token = authorization.split(" ", 1)[1]
    payload = _decode_token(token)
    if not payload or payload.get("kind") != "refresh":
        raise HTTPException(401, "Refresh token non valido")
    user = await store.find_one("users", {"id": payload["sub"]})
    if not user or not user.get("enabled"):
        raise HTTPException(401, "Utente non valido")
    return {"access_token": _make_token(user["id"], ACCESS_TTL, "access")}


@app.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return _public_user(user)


# ── ADMIN — manutenzione ──────────────────────────────────────────────────
@app.post("/admin/backfill_serials")
async def backfill_serials(_: dict = Depends(get_admin_user)):
    """Assegna il seriale progressivo a tutti i ticket che ne sono privi.
    Idempotente: i ticket già serializzati non vengono toccati.
    Utile dopo il primo deploy della funzionalità seriali."""
    result = await _migrate_ticket_serials()
    return result


# ── USERS (admin only) ────────────────────────────────────────────────────
@app.get("/users")
async def list_users(_: dict = Depends(get_admin_user)):
    users = await store.find("users")
    return [_public_user(u) for u in users]


@app.patch("/users/{user_id}")
async def update_user(user_id: str, req: UpdateUserReq, _: dict = Depends(get_admin_user)):
    # exclude_unset=True: includi solo i campi presenti nel body JSON, anche se None.
    # Questo permette di "azzerare" esplicitamente un campo (es. sold_to_code=None
    # per scollegare un cliente). Il vecchio filtro `if v is not None` lo bloccava.
    patch = req.model_dump(exclude_unset=True)
    out = await store.update_one("users", {"id": user_id}, patch)
    if not out:
        raise HTTPException(404, "Utente non trovato")
    return _public_user(out)


@app.delete("/users/{user_id}")
async def delete_user(user_id: str, current: dict = Depends(get_admin_user)):
    if user_id == current["id"]:
        raise HTTPException(400, "Non puoi cancellare te stesso")
    ok = await store.delete_one("users", {"id": user_id})
    if not ok:
        raise HTTPException(404, "Utente non trovato")
    return {"ok": True}


# ── CLIENTI ───────────────────────────────────────────────────────────────
@app.get("/clienti")
async def list_clienti(user: dict = Depends(get_current_user)):
    """Admin vede tutti; customer vede solo se stesso (per coerenza)."""
    all_clienti = await store.find("clienti")
    if user.get("role") == "admin":
        return all_clienti
    return [c for c in all_clienti if c["sold_to"] == user.get("sold_to_code")]


# ── MAGAZZINI ─────────────────────────────────────────────────────────────
@app.get("/magazzini")
async def list_magazzini(user: dict = Depends(get_current_user)):
    return await store.find("magazzini")


# ── ORDINI ────────────────────────────────────────────────────────────────
# IMPORTANTE: l'ordine di dichiarazione di queste route conta!
# Le route con path FISSO (/ordini, /ordini/import, /ordini/import_master)
# DEVONO essere dichiarate PRIMA di quelle con path PARAMETRIZZATO
# (/ordini/{sap_order_number}). Altrimenti FastAPI matcha "import_master"
# come fosse un sap_order_number e ritorna 405 sui POST.
@app.get("/ordini")
async def list_ordini(user: dict = Depends(get_current_user)):
    """Admin vede tutti, customer vede solo quelli del proprio sold_to."""
    all_orders = await store.find("ordini")
    if user.get("role") == "customer":
        all_orders = [o for o in all_orders if o.get("sold_to") == user.get("sold_to_code")]
    return [_enrich_order(o) for o in all_orders]


# Route POST con path fisso PRIMA di quella parametrizzata
@app.post("/ordini/import_master")
async def import_master_excel(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Import master ottimizzato: parser read-only (5x più veloce) + bulk_write
    su Mongo (1 batch invece di 133+ chiamate seriali). Sotto i 100s di
    timeout di Render free anche con 18k righe.

    Logica:
    - clienti: match case-insensitive, fusione automatica "Esposito" ↔ "ESPOSITO"
    - magazzini: zones mergiate
    - ordini: preserva arrived_qty / picked_up_qty per EAN match
    - users: ricollegamento case-fix automatico
    """
    import time
    t0 = time.time()
    from master_parser import parse_master
    content = await file.read()
    tmp = DATA_DIR / f"_master_{uuid.uuid4().hex[:8]}.xlsx"
    tmp.write_bytes(content)
    try:
        data = parse_master(str(tmp))
    except Exception as e:
        tmp.unlink(missing_ok=True)
        raise HTTPException(400, f"Errore parsing master: {e}")
    tmp.unlink(missing_ok=True)
    t_parse = time.time() - t0

    report = {
        "ordini_nuovi": 0,
        "users_ricollegati": 0,
        "tempo_parsing_s": round(t_parse, 1),
        "tempo_totale_s": 0,
    }

    # ── WIPE TOTALE ──────────────────────────────────────────────────────
    # Cancello completamente tutte le entità derivate dal master prima di
    # reinserirle. Foto articoli sono in collezione separata e NON vengono
    # toccate (sono caricate manualmente dall'admin, vivono di vita propria).
    report["wipe_clienti"]       = await store.delete_many("clienti")
    report["wipe_magazzini"]     = await store.delete_many("magazzini")
    report["wipe_ordini"]        = await store.delete_many("ordini")
    report["wipe_carichi"]       = await store.delete_many("carichi")
    report["wipe_ritiri"]        = await store.delete_many("ritiri")
    report["wipe_orfani_carichi"] = await store.delete_many("orfani_carichi")
    report["wipe_orfani_ritiri"]  = await store.delete_many("orfani_ritiri")

    # ── REINSERT ─────────────────────────────────────────────────────────
    # Clienti: insert lineare, niente dedup necessario (collection vuota).
    if data["clienti"]:
        await store.bulk_upsert("clienti", data["clienti"], "sold_to")
    report["clienti_nuovi"] = len(data["clienti"])

    # Magazzini: serializzo le zone (set → list) come fa già il parser, poi insert.
    magazzini_da_inserire = []
    for m in data["magazzini"]:
        magazzini_da_inserire.append({**m, "zones": sorted(set(m.get("zones", [])))})
    if magazzini_da_inserire:
        await store.bulk_upsert("magazzini", magazzini_da_inserire, "id")
    report["magazzini_nuovi"] = len(magazzini_da_inserire)

    # Ordini: insert con chiave order_id (sap::nome). arrived_qty/picked_up_qty
    # partono a 0 perché tutti gli eventi sono stati wipeati.
    if data["ordini"]:
        await store.bulk_upsert("ordini", data["ordini"], "order_id")
    report["ordini_nuovi"] = len(data["ordini"])

    # ── Riallineo users.sold_to_code ai clienti del nuovo master ─────────
    # L'utente è collegato al cliente tramite `sold_to_code` che deve
    # matchare `cliente.sold_to` (derivato dalla col A del master). Se il
    # case cambia (es. "Coba" → "COBA") aggiorno il collegamento mantenendo
    # l'utente. Se la nuova col A non contiene proprio quel cliente
    # (rinominato o rimosso), segnalo l'utente come orfano nel report.
    all_users = await store.find("users")
    all_clienti_after = await store.find("clienti")
    cli_upper_to_actual = {c["sold_to"].upper(): c["sold_to"] for c in all_clienti_after}
    users_orfani: List[dict] = []
    for u in all_users:
        st = u.get("sold_to_code")
        if not st:
            continue
        canonical = cli_upper_to_actual.get(st.upper())
        if canonical:
            if canonical != st:
                await store.update_one("users", {"id": u["id"]}, {"sold_to_code": canonical})
                report["users_ricollegati"] += 1
        else:
            # L'utente punta a un cliente che non c'è più nel master corrente.
            # Non lo cancello (potrebbe essere ancora valido se si reimporta
            # un master che lo riporta), ma lo segnalo all'admin.
            users_orfani.append({
                "id":       u.get("id"),
                "username": u.get("username"),
                "email":    u.get("email"),
                "sold_to_code": st,
            })
    report["users_orfani_count"] = len(users_orfani)
    report["users_orfani"] = users_orfani[:50]

    # Espongo le info diagnostiche del parser (righe lette, scartate, motivi).
    # Le prime 50 di "skipped" bastano per scoprire i pattern senza appesantire
    # la risposta — sono raggruppate per ragione e SAP nel UI.
    report["parser_stats"] = data.get("stats", {})
    report["scarti_count"] = len(data.get("skipped", []))
    report["scarti_dettaglio"] = data.get("skipped", [])[:50]

    report["tempo_totale_s"] = round(time.time() - t0, 1)
    return {"ok": True, **report}


@app.post("/ordini/append_master")
async def append_master_excel(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Import incrementale (no-wipe) di un file in formato master.
    A differenza di /ordini/import_master, NON cancella nulla: aggiunge nuovi
    ordini e aggiorna quelli esistenti, preservando arrived_qty/picked_up_qty
    già accumulati dai carichi/ritiri registrati.

    Logica:
    - parser master identico a quello full-import (stesso schema a 42 colonne)
    - clienti: insert se nuovi, lasciati invariati se già esistono (idem per
      magazzini, così non perdo eventuali correzioni manuali al dato)
    - ordini: upsert per order_id (sap::nome). Se esistente, sui SAP collisione
      preservo le quantità arrivate/ritirate per ogni EAN che ricompare.
    - carichi/ritiri/orfani: NON toccati.
    """
    import time
    t0 = time.time()
    from master_parser import parse_master
    content = await file.read()
    tmp = DATA_DIR / f"_append_{uuid.uuid4().hex[:8]}.xlsx"
    tmp.write_bytes(content)
    try:
        data = parse_master(str(tmp))
    except Exception as e:
        tmp.unlink(missing_ok=True)
        raise HTTPException(400, f"Errore parsing master: {e}")
    tmp.unlink(missing_ok=True)
    t_parse = time.time() - t0

    report = {
        "tempo_parsing_s": round(t_parse, 1),
        "clienti_nuovi": 0,
        "magazzini_nuovi": 0,
        "ordini_nuovi": 0,
        "ordini_aggiornati": 0,
        "righe_preservate": 0,  # quante righe esistenti hanno mantenuto arrived/picked
    }

    # ── Clienti: insert se non già presenti (match case-insensitive su sold_to)
    existing_clienti = await store.find("clienti")
    clienti_upper = {c["sold_to"].upper() for c in existing_clienti}
    for c in data["clienti"]:
        if c["sold_to"].upper() not in clienti_upper:
            await store.insert("clienti", c)
            clienti_upper.add(c["sold_to"].upper())
            report["clienti_nuovi"] += 1

    # ── Magazzini: insert se id non già presente. Se esiste, fondo le zones
    #    (così non sovrascrivo eventuali zone aggiunte da import precedenti).
    existing_mag = {m["id"]: m for m in await store.find("magazzini")}
    for m in data["magazzini"]:
        zones_new = sorted(set(m.get("zones", [])))
        if m["id"] in existing_mag:
            old = existing_mag[m["id"]]
            merged_zones = sorted(set((old.get("zones") or []) + zones_new))
            if merged_zones != (old.get("zones") or []):
                await store.update_one("magazzini", {"id": m["id"]},
                                       {"zones": merged_zones})
        else:
            await store.insert("magazzini", {**m, "zones": zones_new})
            report["magazzini_nuovi"] += 1

    # ── Ordini: upsert per order_id, preservando arrived/picked per EAN
    for o in data["ordini"]:
        existing = await store.find_one("ordini", {"order_id": o["order_id"]})
        if existing:
            # Mappo le righe esistenti per EAN così conservo le qty già arrivate/ritirate.
            # Caso EAN vuoto: identifico la riga con (article_id, size) come fallback.
            def _row_key(r):
                ean = (r.get("ean") or "").strip()
                if ean:
                    return ("ean", ean)
                return ("art", r.get("article_id", ""), r.get("size", ""))
            old_map = {_row_key(r): r for r in existing.get("righe", [])}
            preserved = 0
            for r in o["righe"]:
                old = old_map.get(_row_key(r))
                if old:
                    if (old.get("arrived_qty") or 0) > 0 or (old.get("picked_up_qty") or 0) > 0:
                        preserved += 1
                    r["arrived_qty"] = old.get("arrived_qty", 0)
                    r["picked_up_qty"] = old.get("picked_up_qty", 0)
            report["righe_preservate"] += preserved
            await store.update_one("ordini", {"order_id": o["order_id"]}, o)
            report["ordini_aggiornati"] += 1
        else:
            await store.insert("ordini", o)
            report["ordini_nuovi"] += 1

    # Espongo info diagnostiche del parser (utile per capire scarti su file piccoli)
    report["parser_stats"] = data.get("stats", {})
    report["scarti_count"] = len(data.get("skipped", []))
    report["scarti_dettaglio"] = data.get("skipped", [])[:50]

    report["tempo_totale_s"] = round(time.time() - t0, 1)
    return {"ok": True, **report}


@app.post("/ordini/import")
async def import_ordine_excel_legacy(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Import ordine singolo (formato Ordine Prova). Legacy: di solito si usa import_master."""
    try:
        import io
        content = await file.read()
        # Riproduco la logica di seed_from_excel.py inline
        from seed_from_excel import parse
        tmp = DATA_DIR / f"_import_{uuid.uuid4().hex[:8]}.xlsx"
        tmp.write_bytes(content)
        try:
            data = parse(str(tmp))
        finally:
            tmp.unlink(missing_ok=True)
    except Exception as e:
        raise HTTPException(400, f"Impossibile leggere Excel: {e}")
    n_new = 0
    n_upd = 0
    for c in data["clienti"]:
        if not await store.find_one("clienti", {"sold_to": c["sold_to"]}):
            await store.insert("clienti", c)
    for m in data["magazzini"]:
        if not await store.find_one("magazzini", {"id": m["id"]}):
            await store.insert("magazzini", m)
    for o in data["ordini"]:
        # Legacy seed_from_excel.py non genera order_id; lo costruisco al volo
        # per back-compat con import singolo.
        if not o.get("order_id"):
            o["order_id"] = f"{o['sap_order_number']}::{o.get('name','')}"
        existing = await store.find_one("ordini", {"order_id": o["order_id"]})
        if existing:
            ean_map = {r["ean"]: r for r in existing["righe"]}
            for r in o["righe"]:
                old = ean_map.get(r["ean"])
                if old:
                    r["arrived_qty"] = old.get("arrived_qty", 0)
                    r["picked_up_qty"] = old.get("picked_up_qty", 0)
            await store.update_one("ordini", {"order_id": o["order_id"]}, o)
            n_upd += 1
        else:
            await store.insert("ordini", o)
            n_new += 1
    return {"ok": True, "nuovi": n_new, "aggiornati": n_upd}


# Export Excel dei propri ordini (cliente) o di tutti (admin).
# Filtri opzionali via query string per esportare un singolo ordine:
#   - order_id   → match esatto (SAP::nome)
#   - sold_to    → restringe per cliente (uppercase-insensitive)
#   - name       → restringe per nome ordine (col B)
#   - supplier   → restringe per fornitore (col D, uppercase)
# Senza filtri: tutti gli ordini visibili all'utente.
# IMPORTANTE: dichiarata PRIMA della route con path parametrico /ordini/{order_id}
# altrimenti "/ordini/export.xlsx" verrebbe matchata come order_id="export.xlsx".
@app.get("/ordini/export.xlsx")
async def export_ordini_xlsx(
    user: dict = Depends(get_current_user),
    order_id: Optional[str] = None,
    sold_to: Optional[str] = None,
    name: Optional[str] = None,
    supplier: Optional[str] = None,
):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    all_orders = await store.find("ordini")
    if user.get("role") == "customer":
        all_orders = [o for o in all_orders if o.get("sold_to") == user.get("sold_to_code")]

    # Filtri opzionali per export di un singolo ordine (logico o specifico)
    if order_id:
        all_orders = [o for o in all_orders if (o.get("order_id") or "") == order_id]
    if sold_to:
        all_orders = [o for o in all_orders if (o.get("sold_to") or "").upper() == sold_to.upper()]
    if name:
        all_orders = [o for o in all_orders if (o.get("name") or "") == name]
    if supplier:
        sup_norm = supplier.upper()
        def _sup_of(o: dict) -> str:
            return (o.get("fornitore") or "").upper() or "FORNITORE"
        all_orders = [o for o in all_orders if _sup_of(o) == sup_norm]

    if not all_orders:
        # Non sollevo 404: ritorno comunque un file vuoto (header) così l'admin
        # capisce di non aver match invece di vedere un errore opaco.
        pass

    # ── Indice DDT per EAN: {order_id: {ean: "N024/26 (20pz) · N025/26 (10pz)"}} ──
    all_ritiri = await store.find("ritiri")
    ddt_by_ean: dict = {}   # key: (order_id, ean) → lista di (ddt_ref, qta, data_uscita)
    for rit in all_ritiri:
        oid = rit.get("order_id", "")
        for rr in rit.get("righe", []):
            dref = rr.get("ddt_ref", "")
            if not dref:
                dref = rit.get("ddt_ref", "")   # fallback al livello evento
            if not dref:
                continue
            key = (oid, str(rr.get("ean", "")))
            ddt_by_ean.setdefault(key, []).append(
                (dref, int(rr.get("qta", 0) or 0), rr.get("data_uscita", "") or rit.get("data_ritiro", ""))
            )

    def _ddt_str(order_id, ean):
        entries = ddt_by_ean.get((order_id, str(ean)), [])
        if not entries:
            return ""
        parts = [f"{d} ({q}pz)" for d, q, _ in entries]
        return " · ".join(parts)

    def _ddt_date_str(order_id, ean):
        entries = ddt_by_ean.get((order_id, str(ean)), [])
        if not entries:
            return ""
        dates = sorted({dt for _, _, dt in entries if dt})
        return " · ".join(dates)

    wb = Workbook()
    # ── Foglio 1: Dettaglio (una riga per ogni SKU+taglia) ────────────────
    ws = wb.active
    ws.title = "Dettaglio"
    headers_det = [
        "Cliente", "Nome ordine", "Fornitore", "Consegna",
        "Articolo", "Descrizione", "Colore", "Taglia", "EAN",
        "Prezzo LP", "Sconto %",
        "Attesi", "Arrivati", "Ritirati",
        "DDT Uscita", "Data Uscita",
    ]
    ws.append(headers_det)
    # Stile header
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    def _consegna(o, r):
        kind = (r.get("delivery_kind") or o.get("delivery_kind") or "").strip()
        val  = (r.get("delivery_value") or o.get("delivery_value") or "").strip()
        if kind in ("week", "month") and val:
            return val
        if kind == "none":
            return "Da confermare"
        return o.get("settimana_consegna", "")

    for o in sorted(all_orders, key=lambda x: (x.get("sold_to",""), x.get("name",""))):
        oid = o.get("order_id", "")
        for r in (o.get("righe", []) or []):
            ean = r.get("ean", "")
            ws.append([
                o.get("sold_to", ""),
                o.get("name", ""),
                o.get("fornitore", ""),
                _consegna(o, r),
                r.get("article_id", ""),
                r.get("article_name", ""),
                r.get("color", ""),
                r.get("size", ""),
                ean,
                round(float(r.get("lp", 0) or 0), 2),
                round(float(r.get("sconto", 0) or 0), 2),
                int(r.get("pezzi_confermati", 0) or 0),
                int(r.get("arrived_qty", 0) or 0),
                int(r.get("picked_up_qty", 0) or 0),
                _ddt_str(oid, ean),
                _ddt_date_str(oid, ean),
            ])
    # Auto-width approssimativo (limitato per non esagerare)
    for col_idx, h in enumerate(headers_det, 1):
        max_len = max([len(str(h))] + [len(str(ws.cell(r, col_idx).value or "")) for r in range(2, ws.max_row + 1)])
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(max_len + 2, 10), 40)
    ws.freeze_panes = "A2"

    # ── Foglio 2: Sommario (una riga per (Cliente, Nome ordine, Fornitore)) ─
    ws2 = wb.create_sheet("Sommario")
    headers_sum = ["Cliente", "Nome ordine", "Fornitore", "Consegna",
                   "Attesi", "Arrivati", "Ritirati", "In magazzino"]
    ws2.append(headers_sum)
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    summary: dict = {}
    for o in all_orders:
        key = (o.get("sold_to", ""), o.get("name", ""), o.get("fornitore", ""))
        if key not in summary:
            summary[key] = {"consegna": "", "exp": 0, "arr": 0, "pck": 0}
        s = summary[key]
        for r in (o.get("righe", []) or []):
            s["exp"] += int(r.get("pezzi_confermati", 0) or 0)
            s["arr"] += int(r.get("arrived_qty", 0) or 0)
            s["pck"] += int(r.get("picked_up_qty", 0) or 0)
            if not s["consegna"]:
                s["consegna"] = _consegna(o, r)
    for key in sorted(summary.keys()):
        s = summary[key]
        ws2.append([key[0], key[1], key[2], s["consegna"],
                    s["exp"], s["arr"], s["pck"], s["arr"] - s["pck"]])
    for col_idx, h in enumerate(headers_sum, 1):
        max_len = max([len(str(h))] + [len(str(ws2.cell(r, col_idx).value or "")) for r in range(2, ws2.max_row + 1)])
        ws2.column_dimensions[ws2.cell(1, col_idx).column_letter].width = min(max(max_len + 2, 10), 40)
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Nome file (data SEMPRE in coda, formato gg.mm.aaaa):
    #   - singolo ordine logico       → "<nome_ordine>_<gg.mm.aaaa>.xlsx"
    #   - singolo ordine + fornitore  → "<nome_ordine>_<fornitore>_<gg.mm.aaaa>.xlsx"
    #   - tutti gli ordini di un cli  → "ordini-<cliente>_<gg.mm.aaaa>.xlsx"
    # Sanitizzo i caratteri illegali (/ \ : * ? " < > |).
    def _safe(s: str) -> str:
        return re.sub(r"[\\/:*?\"<>|]", "_", (s or "").strip())
    now = datetime.now(timezone.utc)
    date_it = f"{now.day:02d}.{now.month:02d}.{now.year}"
    if name:
        if supplier:
            fname = f"{_safe(name)}_{_safe(supplier)}_{date_it}.xlsx"
        else:
            fname = f"{_safe(name)}_{date_it}.xlsx"
    else:
        who = sold_to or user.get("sold_to_code") or "tutti"
        fname = f"ordini-{_safe(who)}_{date_it}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# Route con PATH PARAMETRIZZATO — dichiarate DOPO quelle con path fisso
# IMPORTANTE: il path param è ora "order_id" (formato "sap::nome"), non più
# il SAP secco — lo stesso SAP può corrispondere a più ordini distinti.
@app.get("/ordini/{order_id:path}")
async def get_ordine(order_id: str, user: dict = Depends(get_current_user)):
    o = await store.find_one("ordini", {"order_id": order_id})
    if not o:
        raise HTTPException(404, "Ordine non trovato")
    if user.get("role") == "customer" and o.get("sold_to") != user.get("sold_to_code"):
        raise HTTPException(403, "Non puoi vedere ordini di altri clienti")
    return _enrich_order(o)


@app.get("/_debug/ordini_per_nome")
async def debug_ordini_per_nome(q: str, _: dict = Depends(get_admin_user)):
    """Diagnostica admin: dato un frammento di nome ordine, ritorna TUTTI i SAP
    in DB che contengono quel frammento, raggruppati per (sold_to, nome_norm).
    Serve per scoprire perché il cliente "non vede" certi pezzi: di solito è
    una variazione di col B (spazio, suffisso, ecc.) che spezza l'aggregazione.
    """
    qn = (q or "").strip().lower()
    if not qn:
        raise HTTPException(400, "q richiesto")
    all_orders = await store.find("ordini")
    matches = []
    for o in all_orders:
        name = (o.get("name") or "")
        if qn in name.lower():
            matches.append({
                "sap_order_number": o.get("sap_order_number"),
                "sold_to":      o.get("sold_to"),
                "name":         name,
                "name_repr":    repr(name),  # mostra spazi/caratteri invisibili
                "fornitore":    o.get("fornitore", ""),
                "personal_reference": o.get("personal_reference", ""),
                "warehouse_id": o.get("warehouse_id"),
                "settimana_consegna": o.get("settimana_consegna", ""),
                "delivery_kind":  o.get("delivery_kind", ""),
                "delivery_value": o.get("delivery_value", ""),
                "righe_count": len(o.get("righe", [])),
                "total_expected": sum(r.get("expected_qty", 0) for r in o.get("righe", [])),
                "total_arrived":  sum(r.get("arrived_qty",  0) for r in o.get("righe", [])),
                "total_picked_up": sum(r.get("picked_up_qty", 0) for r in o.get("righe", [])),
            })
    # Raggruppo per (sold_to, name_norm) per evidenziare gli split
    by_key: dict = {}
    for m in matches:
        k = f"{m['sold_to']}::{(m['name'] or '').strip()}"
        by_key.setdefault(k, []).append(m)
    summary = [
        {
            "key": k,
            "n_saps": len(v),
            "tot_pieces": sum(x["total_expected"] for x in v),
            "saps": v,
        }
        for k, v in by_key.items()
    ]
    summary.sort(key=lambda x: -x["tot_pieces"])
    return {
        "query": q,
        "match_totali": len(matches),
        "gruppi": summary,
    }


class MoveOrdineReq(BaseModel):
    new_name: str = Field(..., min_length=1, max_length=200)


@app.patch("/ordini/{order_id:path}/move")
async def move_ordine(order_id: str, req: MoveOrdineReq,
                       _: dict = Depends(get_admin_user)):
    """Sposta un ordine (singolo SAP) sotto un altro nome ordine cliente.
    Esempio: SAP 4712978667 attualmente sotto '26-03175' va riassegnato a '26-03172'.

    Cambia il campo `name` e ricalcola `order_id = sap::new_name`. Aggiorna
    anche i carichi/ritiri che puntano al vecchio order_id (per preservare la
    storia degli arrivi/ritiri). Le righe articolo/qty rimangono identiche.

    Errori:
    - 404 se l'ordine non esiste
    - 409 se esiste già un altro ordine con stesso SAP e nuovo nome
      (in quel caso meglio cancellare uno dei due o usare append_master)
    """
    ordine = await store.find_one("ordini", {"order_id": order_id})
    if not ordine:
        raise HTTPException(404, "Ordine non trovato")

    new_name = (req.new_name or "").strip()
    if not new_name:
        raise HTTPException(400, "Nuovo nome ordine vuoto")
    if new_name == ordine.get("name"):
        raise HTTPException(400, "Il nuovo nome è uguale a quello attuale")

    sap = ordine.get("sap_order_number") or ""
    if not sap:
        raise HTTPException(400, "Ordine senza SAP, impossibile ricalcolare order_id")
    new_order_id = f"{sap}::{new_name}"

    # Conflitto: esiste già un ordine con stesso SAP + nuovo nome?
    collision = await store.find_one("ordini", {"order_id": new_order_id})
    if collision:
        raise HTTPException(409,
            f"Esiste già un ordine '{new_name}' con SAP {sap}. "
            f"Cancellalo prima oppure usa 'Aggiungi ordini (master)' per fonderli.")

    # Update ordine: cambio name e order_id
    await store.update_one("ordini", {"order_id": order_id},
                           {"name": new_name, "order_id": new_order_id})

    # Update carichi linkati: bridgio su nuovo order_id
    carichi_aggiornati = 0
    for c in (await store.find("carichi", {"order_id": order_id})):
        await store.update_one("carichi", {"id": c["id"]},
                               {"order_id": new_order_id})
        carichi_aggiornati += 1

    # Update ritiri linkati
    ritiri_aggiornati = 0
    for r in (await store.find("ritiri", {"order_id": order_id})):
        await store.update_one("ritiri", {"id": r["id"]},
                               {"order_id": new_order_id})
        ritiri_aggiornati += 1

    return {
        "ok": True,
        "old_order_id": order_id,
        "new_order_id": new_order_id,
        "old_name": ordine.get("name"),
        "new_name": new_name,
        "sap_order_number": sap,
        "carichi_aggiornati": carichi_aggiornati,
        "ritiri_aggiornati": ritiri_aggiornati,
    }


class PatchExpectedQtyReq(BaseModel):
    expected_qty: int


@app.patch("/ordini/{order_id}/righe/{ean}/expected_qty")
async def patch_expected_qty(
    order_id: str,
    ean: str,
    req: PatchExpectedQtyReq,
    _: dict = Depends(get_admin_user),
):
    """Modifica manuale la qty attesa di una riga ordine (admin only).
    Sovrascrive pezzi_confermati così _enrich_order usa il nuovo valore."""
    if req.expected_qty < 0:
        raise HTTPException(400, "expected_qty non può essere negativo")
    ordine = await store.find_one("ordini", {"order_id": order_id})
    if not ordine:
        raise HTTPException(404, "Ordine non trovato")
    riga = next((r for r in ordine.get("righe", []) if r.get("ean") == ean), None)
    if not riga:
        raise HTTPException(404, f"Riga con EAN {ean} non trovata in questo ordine")
    riga["pezzi_confermati"] = req.expected_qty
    await store.update_one("ordini", {"order_id": order_id}, {"righe": ordine["righe"]})
    return {"ok": True, "ean": ean, "expected_qty": req.expected_qty}


@app.delete("/ordini/{order_id:path}")
async def delete_ordine(order_id: str, _: dict = Depends(get_admin_user)):
    """Cancella un ordine e tutti gli eventi (carichi/ritiri) ad esso collegati.
    Solo admin. Utile per ripulire ordini obsoleti che non sono più nel master."""
    ok = await store.delete_one("ordini", {"order_id": order_id})
    if not ok:
        raise HTTPException(404, "Ordine non trovato")
    # Pulizia eventi: i carichi/ritiri sono linkati per order_id
    n_car = 0
    n_rit = 0
    for c in (await store.find("carichi", {"order_id": order_id})):
        if await store.delete_one("carichi", {"id": c["id"]}):
            n_car += 1
    for r in (await store.find("ritiri", {"order_id": order_id})):
        if await store.delete_one("ritiri", {"id": r["id"]}):
            n_rit += 1
    return {"ok": True, "ordine_cancellato": order_id,
            "carichi_cancellati": n_car, "ritiri_cancellati": n_rit}


# ── Helper: magazzino a consegna automatica ──────────────────────────────
async def _is_autodelivery_warehouse(magazzino_id: str) -> bool:
    """Ritorna True per magazzini-cliente dove l'arrivo equivale automaticamente
    a consegna al cliente (es. Bussolengo: il magazzino è del cliente stesso)."""
    if not magazzino_id:
        return False
    mag = await store.find_one("magazzini", {"id": magazzino_id})
    if not mag:
        return False
    return mag.get("city", "").upper() == "BUSSOLENGO"


# ── CARICHI MAGAZZINO (admin) ─────────────────────────────────────────────
@app.post("/carichi")
async def crea_carico(req: CaricoReq, _: dict = Depends(get_admin_user)):
    """Registra un carico merce: incrementa arrived_qty sulle righe dell'ordine.
    Per i magazzini a consegna automatica (es. Bussolengo) aggiorna anche picked_up_qty."""
    ordine = await store.find_one("ordini", {"order_id": req.order_id})
    if not ordine:
        raise HTTPException(404, "Ordine non trovato")
    is_auto = await _is_autodelivery_warehouse(req.magazzino_id)
    # Aggiorna righe ordine
    righe_aggiornate = []
    for upd in req.righe:
        row = next((r for r in ordine["righe"] if r["ean"] == upd.ean), None)
        if not row:
            raise HTTPException(400, f"EAN {upd.ean} non presente nell'ordine")
        expected = max(row.get("total_qty", 0) - row.get("rejected_qty", 0), 0)
        new_arrived = row.get("arrived_qty", 0) + upd.qta
        if new_arrived > expected:
            raise HTTPException(400, f"EAN {upd.ean}: carico {new_arrived} supera atteso {expected}")
        row["arrived_qty"] = new_arrived
        # Bussolengo (magazzino cliente): arrivo = consegna automatica
        if is_auto:
            row["picked_up_qty"] = row.get("picked_up_qty", 0) + upd.qta
        righe_aggiornate.append({"ean": upd.ean, "qta": upd.qta})
    await store.update_one("ordini", {"order_id": req.order_id},
                            {"righe": ordine["righe"]})
    # Salva evento (mantengo entrambi i campi: order_id per back-link,
    # sap_order_number per riferimento di sistema)
    carico = {
        "id": str(uuid.uuid4()),
        "order_id": req.order_id,
        "sap_order_number": ordine.get("sap_order_number", ""),
        "magazzino_id": req.magazzino_id,
        "data_carico": req.data_carico or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "note": req.note or "",
        "righe": righe_aggiornate,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await store.insert("carichi", carico)
    return carico


@app.get("/carichi")
async def list_carichi(user: dict = Depends(get_admin_user)):
    return await store.find("carichi")


@app.get("/carichi/sessioni")
async def list_carichi_sessioni(_: dict = Depends(get_admin_user)):
    """Ritorna i carichi raggruppati per sessione di import (un import = una riga)."""
    carichi = await store.find("carichi")
    groups: dict = {}
    for c in carichi:
        note = c.get("note", "")
        fname = c.get("filename") or (
            note.replace("Import file: ", "") if "Import file:" in note else note
        )
        if c.get("session_id"):
            sid = c["session_id"]
        else:
            # Vecchi carichi senza session_id: raggruppa per (data + nomefile)
            # così tutti gli ordini dello stesso import appaiono come una riga sola.
            sid = f"legacy::{c.get('data_carico', '')}::{fname}"
        if sid not in groups:
            groups[sid] = {
                "session_id": sid,
                "data_carico": c.get("data_carico", ""),
                "filename": fname,
                "is_pluri": c.get("is_pluri", False),
                "carichi": [],
                "created_at": "",
            }
        groups[sid]["carichi"].append(c)
        if c.get("created_at", "") > groups[sid]["created_at"]:
            groups[sid]["created_at"] = c.get("created_at", "")

    result = []
    for g in groups.values():
        tot_pz = sum(sum(r.get("qta", 0) for r in c.get("righe", [])) for c in g["carichi"])
        n_righe = sum(len(c.get("righe", [])) for c in g["carichi"])
        result.append({
            "session_id": g["session_id"],
            "data_carico": g["data_carico"],
            "filename": g["filename"],
            "is_pluri": g["is_pluri"],
            "n_ordini": len(g["carichi"]),
            "n_righe": n_righe,
            "tot_pz": tot_pz,
            "carichi": g["carichi"],
            "created_at": g["created_at"],
        })
    result.sort(key=lambda x: x["created_at"], reverse=True)
    return result


@app.delete("/carichi/sessioni/{session_id:path}")
async def delete_carico_sessione(session_id: str, _: dict = Depends(get_admin_user)):
    """Cancella tutti i carichi di una sessione e fa rollback degli arrived_qty."""
    all_carichi = await store.find("carichi")
    if session_id.startswith("legacy::"):
        # Chiave legacy: "legacy::data_carico::filename"
        parts = session_id.split("::", 2)
        data_c = parts[1] if len(parts) > 1 else ""
        fname  = parts[2] if len(parts) > 2 else ""
        def _same_fname(c):
            note = c.get("note", "")
            cf = c.get("filename") or (note.replace("Import file: ", "") if "Import file:" in note else note)
            return cf == fname
        carichi = [c for c in all_carichi
                   if not c.get("session_id") and c.get("data_carico") == data_c and _same_fname(c)]
    else:
        carichi = [c for c in all_carichi if c.get("session_id") == session_id]
    if not carichi:
        raise HTTPException(404, "Sessione non trovata")

    is_pluri_session = any(c.get("is_pluri") for c in carichi)
    warnings: list = []

    for carico in carichi:
        order_id = carico.get("order_id")
        ordine = await store.find_one("ordini", {"order_id": order_id}) if order_id else None
        if ordine:
            if is_pluri_session:
                # Pluri: azzera gli EAN che erano stati importati da questo carico
                eans_in_carico = {r["ean"] for r in carico.get("righe", [])}
                for row in ordine["righe"]:
                    if row["ean"] in eans_in_carico:
                        row["arrived_qty"] = 0
            else:
                for upd in carico.get("righe", []):
                    row = next((r for r in ordine["righe"] if r["ean"] == upd["ean"]), None)
                    if not row:
                        continue
                    curr = row.get("arrived_qty", 0)
                    picked = row.get("picked_up_qty", 0)
                    new_arr = max(curr - upd["qta"], 0)
                    if new_arr < picked:
                        warnings.append(
                            f"EAN {upd['ean']}: rollback bloccato a {picked} pz (già ritirati)"
                        )
                        new_arr = picked
                    row["arrived_qty"] = new_arr
            await store.update_one("ordini", {"order_id": ordine["order_id"]},
                                   {"righe": ordine["righe"]})
        await store.delete_one("carichi", {"id": carico["id"]})

    return {"ok": True, "n_cancellati": len(carichi), "warnings": warnings}


@app.delete("/carichi/{carico_id}")
async def delete_carico(carico_id: str, _: dict = Depends(get_admin_user)):
    """Cancella un evento di carico e ROLLBACK arrived_qty sulle righe ordine.
    Se l'ordine ha già subito ritiri (picked_up_qty > 0) e il rollback porterebbe
    arrived_qty < picked_up_qty, la qty viene comunque scalata fino a 0
    (mai negativa) e viene segnalato nel report.
    """
    carico = await store.find_one("carichi", {"id": carico_id})
    if not carico:
        raise HTTPException(404, "Carico non trovato")
    # Risolvo l'ordine via order_id (nuovo); fallback su sap_order_number per
    # eventi storici creati prima del refactor.
    order_id = carico.get("order_id")
    if order_id:
        ordine = await store.find_one("ordini", {"order_id": order_id})
    else:
        sap_legacy = carico.get("sap_order_number")
        ordine = await store.find_one("ordini", {"sap_order_number": sap_legacy}) if sap_legacy else None
    warnings: list = []
    is_auto = await _is_autodelivery_warehouse(carico.get("magazzino_id", ""))
    if ordine:
        for upd in carico.get("righe", []):
            row = next((r for r in ordine["righe"] if r["ean"] == upd["ean"]), None)
            if not row:
                warnings.append(f"EAN {upd['ean']} non più nell'ordine, skip rollback")
                continue
            curr = row.get("arrived_qty", 0)
            picked = row.get("picked_up_qty", 0)
            delta = min(upd["qta"], curr)          # pezzi effettivamente tolti
            new_arr = max(curr - upd["qta"], 0)
            if is_auto:
                # Bussolengo: rollback speculare anche su picked_up_qty
                row["picked_up_qty"] = max(0, picked - delta)
            elif new_arr < picked:
                warnings.append(
                    f"EAN {upd['ean']}: rollback bloccato a {picked} "
                    f"(già ritirati). Differenza: {picked - new_arr} pz."
                )
                new_arr = picked
            row["arrived_qty"] = new_arr
        await store.update_one("ordini", {"order_id": ordine["order_id"]},
                                {"righe": ordine["righe"]})
    else:
        warnings.append("Ordine non più presente: cancello solo l'evento")
    await store.delete_one("carichi", {"id": carico_id})
    return {"ok": True, "warnings": warnings}


# ── CARICO da file (auto-detect magazzino) ──────────────────────────────────
# Flusso:
# 1. POST /carichi/preview con file → backend auto-rileva magazzino, parsa,
#    ritorna preview (righe match + orfani + conflitti). Non scrive nulla.
# 2. User conferma sull'UI (eventualmente sovrascrivendo il magazzino rilevato)
# 3. POST /carichi/import_commit con file + magazzino_id → backend applica.

async def _preview_carico_logic(file_bytes: bytes, filename: str) -> dict:
    """Logica condivisa di parsing + matching: usata sia da preview che da commit
    (così evitiamo di duplicare codice e garantiamo coerenza di interpretazione)."""
    from warehouse_parser import detect_carico, PLURI_WAREHOUSE
    tmp = DATA_DIR / f"_carico_{uuid.uuid4().hex[:8]}.xlsx"
    tmp.write_bytes(file_bytes)
    try:
        detected = detect_carico(str(tmp))
    finally:
        tmp.unlink(missing_ok=True)
    if not detected:
        raise HTTPException(400, "Formato file non riconosciuto. "
                                 "Nessun parser di magazzino corrisponde a questo file.")
    warehouse_name, righe_parsate = detected

    is_pluri = warehouse_name == PLURI_WAREHOUSE

    if is_pluri:
        # Formato multi-cliente: nessun magazzino fisso nel file.
        # Il magazzino viene scelto dall'admin al momento del commit.
        mag = {"id": None, "name": "File multi-cliente (non-COBA)"}
    else:
        # Lookup magazzino in DB (case-insensitive sul name)
        all_mag = await store.find("magazzini")
        mag = next((m for m in all_mag
                    if m.get("name", "").strip().upper() == warehouse_name.strip().upper()),
                   None)
        if not mag:
            raise HTTPException(404,
                f"Magazzino '{warehouse_name}' rilevato dal file ma non presente "
                f"in anagrafica. Importa prima il master con questo magazzino.")

    # Costruisco indice EAN → (ordine, riga_idx).
    # Per i formati standard filtro solo ordini del magazzino rilevato.
    # Per il formato pluri (multi-cliente) includo tutti gli ordini: il SAP
    # nel file è sufficiente a disambiguare senza filtro per magazzino.
    all_ordini = await store.find("ordini")
    ean_to_match = {}
    for o in all_ordini:
        if not is_pluri:
            if o.get("warehouse_id") and o.get("warehouse_id") != mag["id"]:
                continue
        for ri, r in enumerate(o.get("righe", [])):
            ean_to_match.setdefault(r["ean"], []).append((o, ri))

    # Per il formato pluri il file è cumulativo (aggiornato ogni giorno con nuovi
    # arrivi). Prima di matchare, aggreghiamo per (sap, ean) così otteniamo la
    # quantità TOTALE arrivata secondo il file. Poi calcoliamo il delta rispetto
    # a quanto già registrato nell'app, e usiamo quello come qta da aggiungere.
    if is_pluri:
        agg: dict = {}
        for r in righe_parsate:
            key = (r["sap_order_number"], r["ean"])
            if key not in agg:
                agg[key] = dict(r)
                agg[key]["qta"] = 0
            agg[key]["qta"] += r["qta"]
        righe_da_matchare = list(agg.values())
    else:
        righe_da_matchare = righe_parsate

    righe_match = []
    righe_orphan = []
    for r in righe_da_matchare:
        candidates = ean_to_match.get(r["ean"], [])
        chosen = None
        sap_file = (r.get("sap_order_number") or "").strip()
        # REGOLA: il SAP è DISCRIMINANTE.
        # - Se il file dichiara un SAP, deve corrispondere all'ordine.
        # - Se il SAP del file non matcha nessun candidato, NON faccio fallback
        #   su un ordine diverso con lo stesso EAN: sarebbe un errore grave
        #   (lo stesso articolo può essere su più ordini di clienti diversi).
        #   La riga diventa orfana e l'utente decide.
        # - Solo se il file NON dichiara un SAP, accetto la prima occorrenza.
        if sap_file:
            chosen = next(((o, ri) for o, ri in candidates
                           if str(o["sap_order_number"]).strip() == sap_file), None)
        else:
            if candidates:
                chosen = candidates[0]
        if not chosen:
            # Aggiungo info sul perché alla riga orfana
            r_orph = dict(r)
            if sap_file and candidates:
                r_orph["motivo_orfano"] = (
                    f"SAP {sap_file} dichiarato nel file non trovato fra gli ordini; "
                    f"l'EAN {r['ean']} esiste in altri {len(candidates)} ordini "
                    f"con SAP diverso (non assegnato per sicurezza)"
                )
            elif sap_file:
                r_orph["motivo_orfano"] = f"SAP {sap_file} e EAN {r['ean']} non trovati"
            else:
                r_orph["motivo_orfano"] = f"EAN {r['ean']} non in nessun ordine"
            righe_orphan.append(r_orph)
            continue
        ordine, ri = chosen
        riga = ordine["righe"][ri]
        expected = riga.get("pezzi_confermati",
                            max(riga.get("total_qty", 0) - riga.get("rejected_qty", 0), 0))
        currently_arrived = riga.get("arrived_qty", 0)

        if is_pluri:
            # r["qta"] = totale arrivato nel file (cumulativo).
            # Il file è la BIBBIA: l'ultima versione caricata sovrascrive sempre.
            # delta può essere negativo (rettifica al ribasso).
            file_total = r["qta"]
            delta = file_total - currently_arrived
            new_arrived = file_total
            conflict = None
            if file_total > expected:
                conflict = (
                    f"file giacenza: {file_total} pz arrivati "
                    f"superano atteso ({expected} pz)"
                )
            righe_match.append({
                **r,
                "qta": delta,               # variazione rispetto a prima (può essere negativa)
                "file_total": file_total,   # totale nel file (valore che verrà impostato)
                "ordine_name": ordine.get("name"),
                "ordine_sold_to": ordine.get("sold_to"),
                "ordine_sap": ordine.get("sap_order_number", ""),
                "ordine_id": ordine.get("order_id", ""),
                "expected_qty": expected,
                "currently_arrived": currently_arrived,
                "conflict": conflict,
            })
        else:
            new_arrived = currently_arrived + r["qta"]
            conflict = None
            if new_arrived > expected:
                conflict = f"supera atteso ({new_arrived} > {expected})"
            righe_match.append({
                **r,
                "ordine_name": ordine.get("name"),
                "ordine_sold_to": ordine.get("sold_to"),
                "ordine_sap": ordine.get("sap_order_number", ""),
                # PK reale dell'ordine lato app (formato "sap::nome"); lo uso per
                # raggruppare per ordine nel commit, evitando collisioni quando
                # lo stesso SAP copre più ordini.
                "ordine_id": ordine.get("order_id", ""),
                "expected_qty": expected,
                "currently_arrived": currently_arrived,
                "conflict": conflict,
            })

    # Per il formato pluri gli orfani sono normalmente migliaia (articoli di
    # clienti non registrati nell'app): restituiamo solo i primi 10 per la
    # preview senza appesantire la risposta.
    orphan_detail = righe_orphan[:10] if is_pluri else righe_orphan

    return {
        "magazzino_id": mag["id"],
        "magazzino_name": mag["name"],
        "is_pluri": is_pluri,
        "filename": filename,
        "n_righe": len(righe_parsate),
        "n_match": len(righe_match),
        "n_orphan": len(righe_orphan),
        "n_conflitti": sum(1 for r in righe_match if r.get("conflict")),
        "righe_match": righe_match,
        "righe_orphan": orphan_detail,
    }


@app.post("/carichi/preview")
async def carico_preview(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Anteprima carico: rileva magazzino, parsa, mostra match/orfani.
    NON scrive nulla sul DB. L'utente conferma poi via /carichi/import_commit."""
    content = await file.read()
    return await _preview_carico_logic(content, file.filename or "carico.xlsx")


@app.post("/carichi/import_commit")
async def carico_import_commit(
    file: UploadFile = File(...),
    magazzino_id: str = Form(...),
    _: dict = Depends(get_admin_user),
):
    """Applica il carico letto da file. Il magazzino è passato esplicitamente
    (di solito quello auto-rilevato in preview, ma l'utente può sovrascriverlo).

    Per ogni SAP coinvolto crea un evento `carichi` separato.
    Le righe orfane (EAN non in nessun ordine) finiscono in `orfani_carichi`."""
    content = await file.read()
    preview = await _preview_carico_logic(content, file.filename or "carico.xlsx")

    # Override magazzino_id se diverso dall'auto-detect
    mag = await store.find_one("magazzini", {"id": magazzino_id})
    if not mag:
        raise HTTPException(404, f"Magazzino {magazzino_id} non trovato")
    # Bussolengo (magazzino cliente): arrivo = consegna automatica
    is_auto_delivery = mag.get("city", "").upper() == "BUSSOLENGO"

    now_iso = datetime.now(timezone.utc).isoformat()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    session_id = str(uuid.uuid4())   # raggruppa tutti i carichi di questo import

    # ── Orfani: salva in collezione dedicata, status="pending" ──
    # Per il formato pluri gli orfani sono attesi (migliaia di articoli non in
    # anagrafica): non salviamo per non inquinare orfani_carichi.
    orfani_creati = 0
    if not preview.get("is_pluri"):
        for r in preview["righe_orphan"]:
            await store.insert("orfani_carichi", {
                "id": str(uuid.uuid4()),
                "ean": r["ean"],
                "qta": r["qta"],
                "magazzino_id": magazzino_id,
                "articolo": r.get("articolo", ""),
                "descrizione": r.get("descrizione", ""),
                "mis": r.get("mis", ""),
                "sap_dichiarato": r.get("sap_order_number", ""),
                "doc_forn": r.get("doc_forn", ""),
                "n_ordine_coba": r.get("n_ordine_coba", ""),
                "motivo_orfano": r.get("motivo_orfano", ""),
                "filename": preview["filename"],
                "stato": "pending",
                "created_at": now_iso,
            })
            orfani_creati += 1

    # ── Match: raggruppa per ORDER_ID (non per SAP, lo stesso SAP può
    #     coprire più ordini distinti).
    per_order: dict = {}
    for r in preview["righe_match"]:
        per_order.setdefault(r["ordine_id"], []).append(r)

    eventi_creati = 0
    righe_caricate = 0
    errori: List[str] = []
    for order_id, righe_event in per_order.items():
        ordine = await store.find_one("ordini", {"order_id": order_id})
        if not ordine:
            errori.append(f"Ordine {order_id} non più presente")
            continue
        # Pluri bibbia: EAN non presenti nel file → azzerati (erano arrivati per errore)
        if preview.get("is_pluri"):
            eans_in_file = {r["ean"] for r in righe_event}
            for row in ordine["righe"]:
                if row["ean"] not in eans_in_file:
                    row["arrived_qty"] = 0
        righe_evento = []
        for r in righe_event:
            row = next((x for x in ordine["righe"] if x["ean"] == r["ean"]), None)
            if not row:
                errori.append(f"EAN {r['ean']} sparito da ordine {order_id}")
                continue
            if preview.get("is_pluri"):
                # Pluri: il file è autoritativo → imposta il valore esatto
                row["arrived_qty"] = r["file_total"]
            else:
                row["arrived_qty"] = row.get("arrived_qty", 0) + r["qta"]
            # Bussolengo: arrivo = consegna automatica al cliente
            if is_auto_delivery:
                row["picked_up_qty"] = row.get("picked_up_qty", 0) + r["qta"]
            # Per pluri registra il totale file (non il delta) così il log è leggibile
            qta_log = r["file_total"] if preview.get("is_pluri") else r["qta"]
            righe_evento.append({"ean": r["ean"], "qta": qta_log})
            righe_caricate += 1
        await store.update_one("ordini", {"order_id": order_id},
                                {"righe": ordine["righe"]})
        await store.insert("carichi", {
            "id": str(uuid.uuid4()),
            "session_id": session_id,
            "filename": preview["filename"],
            "is_pluri": preview.get("is_pluri", False),
            "order_id": order_id,
            "sap_order_number": ordine.get("sap_order_number", ""),
            "magazzino_id": magazzino_id,
            "data_carico": today,
            "note": f"Import file: {preview['filename']}",
            "righe": righe_evento,
            "created_at": now_iso,
        })
        eventi_creati += 1

    return {
        "ok": True,
        "magazzino": mag["name"],
        "eventi_carico_creati": eventi_creati,
        "righe_caricate": righe_caricate,
        "orfani_creati": orfani_creati,
        "errori": errori,
    }


# ── ORFANI CARICHI ──────────────────────────────────────────────────────────
@app.get("/orfani_carichi")
async def list_orfani_carichi(_: dict = Depends(get_admin_user)):
    """Lista degli EAN orfani non ancora assegnati a un ordine."""
    return await store.find("orfani_carichi")


@app.delete("/orfani_carichi/{orf_id}")
async def delete_orfano_carico(orf_id: str, _: dict = Depends(get_admin_user)):
    """Cancella un orfano (es. è stato risolto manualmente)."""
    ok = await store.delete_one("orfani_carichi", {"id": orf_id})
    if not ok:
        raise HTTPException(404, "Orfano non trovato")
    return {"ok": True}


# ── RITIRI (admin) ────────────────────────────────────────────────────────
@app.post("/ritiri")
async def crea_ritiro(req: RitiroReq, _: dict = Depends(get_admin_user)):
    """Registra un ritiro: incrementa picked_up_qty sulle righe dell'ordine.
    picked_up_qty non può superare arrived_qty (non puoi ritirare merce non arrivata)."""
    ordine = await store.find_one("ordini", {"order_id": req.order_id})
    if not ordine:
        raise HTTPException(404, "Ordine non trovato")
    righe_aggiornate = []
    for upd in req.righe:
        row = next((r for r in ordine["righe"] if r["ean"] == upd.ean), None)
        if not row:
            raise HTTPException(400, f"EAN {upd.ean} non presente nell'ordine")
        arrived = row.get("arrived_qty", 0)
        new_picked = row.get("picked_up_qty", 0) + upd.qta
        if new_picked > arrived:
            raise HTTPException(400,
                f"EAN {upd.ean}: ritiro {new_picked} supera arrivato {arrived}")
        row["picked_up_qty"] = new_picked
        righe_aggiornate.append({"ean": upd.ean, "qta": upd.qta})
    await store.update_one("ordini", {"order_id": req.order_id},
                            {"righe": ordine["righe"]})
    ritiro = {
        "id": str(uuid.uuid4()),
        "order_id": req.order_id,
        "sap_order_number": ordine.get("sap_order_number", ""),
        "magazzino_id": req.magazzino_id,
        "data_ritiro": req.data_ritiro or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "note": req.note or "",
        "righe": righe_aggiornate,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await store.insert("ritiri", ritiro)
    return ritiro


@app.get("/ritiri")
async def list_ritiri(user: dict = Depends(get_admin_user)):
    return await store.find("ritiri")


@app.delete("/ritiri/{ritiro_id}")
async def delete_ritiro(ritiro_id: str, _: dict = Depends(get_admin_user)):
    """Cancella un evento di ritiro e ROLLBACK picked_up_qty sulle righe ordine."""
    ritiro = await store.find_one("ritiri", {"id": ritiro_id})
    if not ritiro:
        raise HTTPException(404, "Ritiro non trovato")
    order_id = ritiro.get("order_id")
    if order_id:
        ordine = await store.find_one("ordini", {"order_id": order_id})
    else:
        sap_legacy = ritiro.get("sap_order_number")
        ordine = await store.find_one("ordini", {"sap_order_number": sap_legacy}) if sap_legacy else None
    warnings: list = []
    if ordine:
        for upd in ritiro.get("righe", []):
            row = next((r for r in ordine["righe"] if r["ean"] == upd["ean"]), None)
            if not row:
                warnings.append(f"EAN {upd['ean']} non più nell'ordine, skip rollback")
                continue
            row["picked_up_qty"] = max(row.get("picked_up_qty", 0) - upd["qta"], 0)
        await store.update_one("ordini", {"order_id": ordine["order_id"]},
                                {"righe": ordine["righe"]})
    else:
        warnings.append("Ordine non più presente: cancello solo l'evento")
    await store.delete_one("ritiri", {"id": ritiro_id})
    return {"ok": True, "warnings": warnings}


# ── RITIRO da file (auto-detect magazzino, filtra righe gialle) ────────────
# Flusso speculare al carico:
# 1. POST /ritiri/preview → backend rileva magazzino, parsa righe gialle,
#    matcha per (article_id, size). Ritorna preview con match/orfani/ambigui.
# 2. POST /ritiri/import_commit → applica: incrementa picked_up_qty, crea
#    eventi `ritiri`, salva orfani in `orfani_ritiri`.

def _size_norm(s) -> str:
    """Normalizza una taglia per match robusto.
    - lower + strip + sostituisce ½ con .5 / "-" (mezze taglie).
    Es: "10½" == "10.5" == "10-" (alcuni formati Adidas usano "-").
    """
    if s is None:
        return ""
    x = str(s).strip().upper()
    # Mezze taglie: 10½ → 10.5
    x = x.replace("½", ".5")
    # Alcuni master segnano la mezza con "-" alla fine (es. "10-").
    # Normalizziamo "10-" → "10.5" per match con "10½".
    if x.endswith("-") and x[:-1].replace(".", "").isdigit():
        x = x[:-1] + ".5"
    return x


async def _preview_ritiro_logic(file_bytes: bytes, filename: str) -> dict:
    """Logica condivisa preview ritiro: parsing + matching (article_id, size)."""
    from warehouse_parser import detect_ritiro
    tmp = DATA_DIR / f"_ritiro_{uuid.uuid4().hex[:8]}.xlsx"
    tmp.write_bytes(file_bytes)
    try:
        detected = detect_ritiro(str(tmp))
    finally:
        tmp.unlink(missing_ok=True)
    if not detected:
        raise HTTPException(400, "Formato file ritiro non riconosciuto. "
                                 "Nessun parser corrisponde a questo file.")
    warehouse_name, righe_parsate = detected

    all_mag = await store.find("magazzini")
    mag = next((m for m in all_mag
                if m.get("name", "").strip().upper() == warehouse_name.strip().upper()),
               None)
    if not mag:
        raise HTTPException(404,
            f"Magazzino '{warehouse_name}' rilevato dal file ma non presente "
            f"in anagrafica.")

    # Costruisco due indici (filtrati per magazzino):
    # - ean_to_match: EAN → [(ordine, riga_idx)]  (match preciso)
    # - key_to_match: (article_id, size_norm) → [...]  (fallback)
    all_ordini = await store.find("ordini")
    ean_to_match: dict = {}
    key_to_match: dict = {}
    for o in all_ordini:
        if o.get("warehouse_id") and o.get("warehouse_id") != mag["id"]:
            continue
        for ri, r in enumerate(o.get("righe", [])):
            ean = str(r.get("ean", "")).strip()
            if ean:
                ean_to_match.setdefault(ean, []).append((o, ri))
            k = (str(r.get("article_id", "")).strip().upper(),
                 _size_norm(r.get("size")))
            key_to_match.setdefault(k, []).append((o, ri))

    righe_match = []
    righe_orphan = []
    for r in righe_parsate:
        # 1° tentativo: match diretto per EAN se la cella del file è valorizzata
        candidates = []
        matched_by = ""
        ean_file = (r.get("ean") or "").strip()
        if ean_file:
            candidates = ean_to_match.get(ean_file, [])
            matched_by = "ean"
        # 2° tentativo: fallback su (articolo, size)
        if not candidates:
            key = (r["articolo"].strip().upper(), _size_norm(r["size"]))
            candidates = key_to_match.get(key, [])
            matched_by = "articolo+size"
        if not candidates:
            righe_orphan.append(r)
            continue
        # Scegli il candidato con più arrived disponibile (arrived - picked_up)
        def _disp(t):
            _, ri = t
            row = t[0]["righe"][ri]
            return row.get("arrived_qty", 0) - row.get("picked_up_qty", 0)
        candidates_sorted = sorted(candidates, key=_disp, reverse=True)
        ordine, ri = candidates_sorted[0]
        riga = ordine["righe"][ri]
        arrived = riga.get("arrived_qty", 0)
        current_picked = riga.get("picked_up_qty", 0)
        new_picked = current_picked + r["qta"]
        conflict = None
        if new_picked > arrived:
            conflict = f"supera arrivato ({new_picked} > {arrived})"
        ambiguo = None
        if len(candidates) > 1:
            ambiguo = f"{len(candidates)} ordini contengono questo (articolo, size)"
        righe_match.append({
            **r,
            "ordine_name": ordine.get("name"),
            "ordine_sold_to": ordine.get("sold_to"),
            "ordine_sap": ordine.get("sap_order_number", ""),
            "ordine_id": ordine.get("order_id", ""),
            "ean": riga.get("ean"),
            "arrived_qty": arrived,
            "current_picked": current_picked,
            "conflict": conflict,
            "ambiguo": ambiguo,
            "matched_by": matched_by,  # "ean" (preciso) o "articolo+size" (fallback)
        })

    return {
        "magazzino_id": mag["id"],
        "magazzino_name": mag["name"],
        "filename": filename,
        "n_righe": len(righe_parsate),
        "n_match": len(righe_match),
        "n_orphan": len(righe_orphan),
        "n_conflitti": sum(1 for r in righe_match if r.get("conflict")),
        "n_ambigui": sum(1 for r in righe_match if r.get("ambiguo")),
        "n_match_ean": sum(1 for r in righe_match if r.get("matched_by") == "ean"),
        "n_match_fallback": sum(1 for r in righe_match if r.get("matched_by") == "articolo+size"),
        "righe_match": righe_match,
        "righe_orphan": righe_orphan,
    }


@app.post("/ritiri/preview")
async def ritiro_preview(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Anteprima ritiro: rileva magazzino, parsa righe gialle, mostra match."""
    content = await file.read()
    return await _preview_ritiro_logic(content, file.filename or "ritiro.xlsx")


@app.post("/ritiri/import_commit")
async def ritiro_import_commit(
    file: UploadFile = File(...),
    magazzino_id: str = Form(...),
    _: dict = Depends(get_admin_user),
):
    """Applica il ritiro dopo conferma. Le righe orfane (articolo+size non
    trovati in nessun ordine) finiscono in `orfani_ritiri`."""
    content = await file.read()
    preview = await _preview_ritiro_logic(content, file.filename or "ritiro.xlsx")

    mag = await store.find_one("magazzini", {"id": magazzino_id})
    if not mag:
        raise HTTPException(404, f"Magazzino {magazzino_id} non trovato")

    now_iso = datetime.now(timezone.utc).isoformat()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Orfani
    orfani_creati = 0
    for r in preview["righe_orphan"]:
        await store.insert("orfani_ritiri", {
            "id": str(uuid.uuid4()),
            "articolo": r["articolo"],
            "size": r["size"],
            "qta": r["qta"],
            "magazzino_id": magazzino_id,
            "n_ordine_interno": r.get("n_ordine_interno", ""),
            "delivery": r.get("delivery", ""),
            "partenza": r.get("partenza", ""),
            "filename": preview["filename"],
            "stato": "pending",
            "created_at": now_iso,
        })
        orfani_creati += 1

    # Match: raggruppo per ORDER_ID (lo stesso SAP può coprire più ordini)
    per_order: dict = {}
    for r in preview["righe_match"]:
        per_order.setdefault(r["ordine_id"], []).append(r)

    eventi_creati = 0
    righe_ritirate = 0
    errori: List[str] = []
    for order_id, righe_event in per_order.items():
        ordine = await store.find_one("ordini", {"order_id": order_id})
        if not ordine:
            errori.append(f"Ordine {order_id} non più presente")
            continue
        righe_evento = []
        for r in righe_event:
            row = next((x for x in ordine["righe"] if x["ean"] == r["ean"]), None)
            if not row:
                errori.append(f"EAN {r['ean']} sparito da ordine {order_id}")
                continue
            arrived = row.get("arrived_qty", 0)
            new_picked = row.get("picked_up_qty", 0) + r["qta"]
            if new_picked > arrived:
                errori.append(
                    f"Ordine {order_id} EAN {r['ean']}: ritiro {new_picked} supera arrivato {arrived}")
            row["picked_up_qty"] = new_picked
            righe_evento.append({"ean": r["ean"], "qta": r["qta"]})
            righe_ritirate += 1
        await store.update_one("ordini", {"order_id": order_id},
                                {"righe": ordine["righe"]})
        await store.insert("ritiri", {
            "id": str(uuid.uuid4()),
            "order_id": order_id,
            "sap_order_number": ordine.get("sap_order_number", ""),
            "magazzino_id": magazzino_id,
            "data_ritiro": today,
            "note": f"Import file: {preview['filename']}",
            "righe": righe_evento,
            "created_at": now_iso,
        })
        eventi_creati += 1

    return {
        "ok": True,
        "magazzino": mag["name"],
        "eventi_ritiro_creati": eventi_creati,
        "righe_ritirate": righe_ritirate,
        "orfani_creati": orfani_creati,
        "errori": errori,
    }


# ── DDT NEXT: import ritiro da file DDT ─────────────────────────────────────
async def _preview_ddt_logic(file_bytes: bytes, filename: str) -> dict:
    """Preview import DDT: match EAN → ordini (senza filtro magazzino)."""
    from warehouse_parser import DDT_NEXT_WAREHOUSE, detect_ritiro
    tmp = DATA_DIR / f"_ddt_{uuid.uuid4().hex[:8]}.xlsx"
    tmp.write_bytes(file_bytes)
    try:
        detected = detect_ritiro(str(tmp))
    finally:
        tmp.unlink(missing_ok=True)
    if not detected or detected[0] != DDT_NEXT_WAREHOUSE:
        raise HTTPException(400, "Formato non riconosciuto come DDT NEXT. "
                                 "Il file deve avere il foglio DETTAGLIO con colonne EAN, "
                                 "Delivery quantity e Rif ns DDT.")
    _, righe_parsate = detected

    # Indice EAN → [(ordine, riga_idx)] su TUTTI gli ordini
    all_ordini = await store.find("ordini")
    ean_to_match: dict = {}
    for o in all_ordini:
        for ri, r in enumerate(o.get("righe", [])):
            ean = str(r.get("ean", "")).strip()
            if ean:
                ean_to_match.setdefault(ean, []).append((o, ri))

    ddt_ref = next((r.get("ddt_ref", "") for r in righe_parsate if r.get("ddt_ref")), "")

    righe_match = []
    righe_orphan = []
    for r in righe_parsate:
        ean = r.get("ean", "").strip()
        candidates = list(ean_to_match.get(ean, []))
        if not candidates:
            righe_orphan.append(r)
            continue
        # Preferisci ordine con SAP matching
        sap_file = r.get("sap_order_number", "").strip().upper()
        if sap_file:
            exact = [c for c in candidates
                     if c[0].get("sap_order_number", "").strip().upper() == sap_file
                     or c[0].get("name", "").strip().upper() == sap_file]
            if exact:
                candidates = exact
        ordine, ri_idx = candidates[0]
        riga = ordine["righe"][ri_idx]
        arrived = riga.get("arrived_qty", 0)
        current_picked = riga.get("picked_up_qty", 0)
        new_picked = current_picked + r["qta"]
        conflict = None
        if new_picked > arrived:
            conflict = f"ritiro {new_picked} supera arrivato ({arrived})"
        righe_match.append({
            **r,
            "ordine_name": ordine.get("name"),
            "ordine_sold_to": ordine.get("sold_to"),
            "ordine_sap": ordine.get("sap_order_number", ""),
            "ordine_id": ordine.get("order_id", ""),
            "ean": riga.get("ean"),
            "arrived_qty": arrived,
            "current_picked": current_picked,
            "new_picked": new_picked,
            "conflict": conflict,
        })

    return {
        "filename": filename,
        "ddt_ref": ddt_ref,
        "n_righe": len(righe_parsate),
        "n_match": len(righe_match),
        "n_orphan": len(righe_orphan),
        "n_conflitti": sum(1 for r in righe_match if r.get("conflict")),
        "righe_match": righe_match,
        "righe_orphan": righe_orphan[:10],
    }


@app.post("/ritiri/ddt_preview")
async def ritiri_ddt_preview(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Anteprima import DDT NEXT: match EAN → ordini."""
    content = await file.read()
    return await _preview_ddt_logic(content, file.filename or "ddt.xlsx")


@app.post("/ritiri/ddt_commit")
async def ritiri_ddt_commit(file: UploadFile = File(...), _: dict = Depends(get_admin_user)):
    """Applica ritiri dal DDT NEXT. Raggruppa per (ordine, data) e crea eventi ritiro."""
    content = await file.read()
    fname = file.filename or "ddt.xlsx"
    preview = await _preview_ddt_logic(content, fname)

    now_iso = datetime.now(timezone.utc).isoformat()
    session_id = str(uuid.uuid4())
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Data di fallback: estraggo dal nome file "DEL DD.MM.YYYY" se presente
    _dm = re.search(r"DEL\s+(\d{2})\.(\d{2})\.(\d{4})", fname, re.IGNORECASE)
    ddt_date_fallback = (f"{_dm.group(3)}-{_dm.group(2)}-{_dm.group(1)}"
                         if _dm else today)

    # Raggruppa per (order_id, data_ritiro)
    per_group: dict = {}
    for r in preview["righe_match"]:
        data_r = r.get("data_ritiro") or ddt_date_fallback
        key = (r["ordine_id"], data_r)
        per_group.setdefault(key, []).append(r)

    eventi_creati = 0
    righe_ritirate = 0
    errori: List[str] = []
    for (order_id, data_ritiro), righe_group in per_group.items():
        ordine = await store.find_one("ordini", {"order_id": order_id})
        if not ordine:
            errori.append(f"Ordine {order_id} non trovato")
            continue
        righe_evento = []
        for r in righe_group:
            row = next((x for x in ordine["righe"] if x["ean"] == r["ean"]), None)
            if not row:
                errori.append(f"EAN {r['ean']} non trovato in ordine {order_id}")
                continue
            arrived = row.get("arrived_qty", 0)
            new_picked = min(row.get("picked_up_qty", 0) + r["qta"], arrived)
            row["picked_up_qty"] = new_picked
            # Salva DDT ref per-riga: tracciabilità di quale DDT ha portato fuori i pz
            righe_evento.append({
                "ean": r["ean"],
                "qta": r["qta"],
                "ddt_ref": r.get("ddt_ref", ""),
                "data_uscita": data_ritiro,
            })
            righe_ritirate += 1
        await store.update_one("ordini", {"order_id": order_id}, {"righe": ordine["righe"]})
        await store.insert("ritiri", {
            "id": str(uuid.uuid4()),
            "session_id": session_id,
            "filename": preview["filename"],
            "ddt_ref": preview.get("ddt_ref", ""),
            "order_id": order_id,
            "sap_order_number": ordine.get("sap_order_number", ""),
            "magazzino_id": "",
            "data_ritiro": data_ritiro,
            "note": f"DDT {preview.get('ddt_ref', '')} · Import: {preview['filename']}",
            "righe": righe_evento,
            "created_at": now_iso,
        })
        eventi_creati += 1

    return {
        "ok": True,
        "ddt_ref": preview.get("ddt_ref", ""),
        "eventi_creati": eventi_creati,
        "righe_ritirate": righe_ritirate,
        "errori": errori,
    }


# ── ORFANI RITIRI ───────────────────────────────────────────────────────────
@app.get("/orfani_ritiri")
async def list_orfani_ritiri(_: dict = Depends(get_admin_user)):
    return await store.find("orfani_ritiri")


@app.delete("/orfani_ritiri/{orf_id}")
async def delete_orfano_ritiro(orf_id: str, _: dict = Depends(get_admin_user)):
    ok = await store.delete_one("orfani_ritiri", {"id": orf_id})
    if not ok:
        raise HTTPException(404, "Orfano non trovato")
    return {"ok": True}


# ── TICKETS (chat di supporto cliente↔admin su ordine/articolo) ──────────
# Modello: ogni ticket è una conversazione threaded tra il cliente e l'admin
# relativa a uno specifico ordine (PK app = "SAP::nome"). Lo scope può essere:
#   - "ordine"   → domanda generica sull'ordine
#   - "articolo" → domanda su un articolo specifico (article_id), con dati
#                  contestuali (article_name, color, ean, size) salvati nel
#                  ticket così non si perdono se l'ordine cambia.
# Stato: open|closed (chiusura solo admin). Messaggi: lista append-only.
# Contatori unread_admin / unread_customer per il badge bottom-nav.

class TicketCreateReq(BaseModel):
    order_id: str
    scope: str = "ordine"  # "ordine" | "articolo"
    article_id: Optional[str] = None
    article_name: Optional[str] = None
    color: Optional[str] = None
    ean: Optional[str] = None
    size: Optional[str] = None
    # Fornitore se il cliente è in vista filtrata su un fornitore specifico
    # (es. ordine multi-fornitore di cui sta guardando la parte NEXT SRL).
    # Solo informativo: il ticket è già agganciato a order_id.
    supplier: Optional[str] = None
    message: str = Field(..., min_length=1, max_length=4000)


class TicketMessageReq(BaseModel):
    message: str = Field(..., min_length=1, max_length=4000)


class TicketStatusReq(BaseModel):
    status: str = Field(..., pattern="^(open|closed)$")


def _ticket_can_access(ticket: dict, user: dict) -> bool:
    """L'admin vede tutto. Il cliente vede solo i ticket del proprio sold_to."""
    if user.get("role") == "admin":
        return True
    return ticket.get("sold_to") == user.get("sold_to_code")


@app.post("/tickets")
async def create_ticket(req: TicketCreateReq,
                         user: dict = Depends(get_current_user)):
    """Apri un nuovo ticket con un primo messaggio. Sia admin che customer
    possono aprire; in pratica lo userà soprattutto il cliente dal dettaglio
    ordine. Customer può aprire solo su ordini del proprio sold_to."""
    ordine = await store.find_one("ordini", {"order_id": req.order_id})
    if not ordine:
        raise HTTPException(404, "Ordine non trovato")
    if user.get("role") == "customer" and ordine.get("sold_to") != user.get("sold_to_code"):
        raise HTTPException(403, "Non sei autorizzato su questo ordine")
    scope = req.scope if req.scope in ("ordine", "articolo") else "ordine"
    if scope == "articolo" and not (req.article_id or req.ean):
        raise HTTPException(400, "Scope articolo richiede article_id o ean")

    now = datetime.now(timezone.utc).isoformat()
    msg_text = (req.message or "").strip()
    if not msg_text:
        raise HTTPException(400, "Messaggio vuoto")

    # Seriale progressivo: 00001, 00002, …
    serial_num = await store.next_serial("tickets")
    serial = str(serial_num).zfill(5)

    is_customer = user.get("role") == "customer"
    ticket = {
        "id": str(uuid.uuid4()),
        "serial": serial,
        "order_id": req.order_id,
        "sap_order_number": ordine.get("sap_order_number", ""),
        "sold_to": ordine.get("sold_to", ""),
        "order_name": ordine.get("name", ""),
        "scope": scope,
        "article_id": req.article_id or "",
        "article_name": req.article_name or "",
        "color": req.color or "",
        "ean": req.ean or "",
        "size": req.size or "",
        "supplier": req.supplier or "",
        "status": "open",
        "created_by": user["id"],
        "created_by_name": user.get("username") or user.get("email", ""),
        "created_by_role": user.get("role", "customer"),
        "created_at": now,
        "updated_at": now,
        "last_message_at": now,
        "closed_at": None,
        "messages": [
            {
                "id": str(uuid.uuid4()),
                "author_id": user["id"],
                "author_role": user.get("role", "customer"),
                "author_name": user.get("username") or user.get("email", ""),
                "text": msg_text,
                "created_at": now,
            }
        ],
        # Il primo messaggio è non letto per la controparte
        "unread_admin": 1 if is_customer else 0,
        "unread_customer": 0 if is_customer else 1,
    }
    await store.insert("tickets", ticket)
    return ticket


@app.get("/tickets")
async def list_tickets(status: Optional[str] = None,
                        order_id: Optional[str] = None,
                        user: dict = Depends(get_current_user)):
    """Lista ticket visibili all'utente. Admin: tutti. Customer: solo i suoi.
    Filtri opzionali: status (open|closed), order_id."""
    q: Dict[str, Any] = {}
    if status in ("open", "closed"):
        q["status"] = status
    if order_id:
        q["order_id"] = order_id
    tickets = await store.find("tickets", q or None)
    if user.get("role") != "admin":
        tickets = [t for t in tickets if t.get("sold_to") == user.get("sold_to_code")]
    # Sort: aperti prima, poi per last_message_at desc
    def _sort_key(t):
        return (0 if t.get("status") == "open" else 1,
                -(t.get("last_message_at") or "" > "") if False else "",
                t.get("last_message_at") or "")
    tickets.sort(key=lambda t: (0 if t.get("status") == "open" else 1,
                                 t.get("last_message_at") or ""),
                  reverse=False)
    # In realtà voglio aperti prima ma per last_message_at desc all'interno
    tickets.sort(key=lambda t: t.get("last_message_at") or "", reverse=True)
    tickets.sort(key=lambda t: 0 if t.get("status") == "open" else 1)
    return tickets


@app.get("/tickets/unread_count")
async def tickets_unread_count(user: dict = Depends(get_current_user)):
    """Numero messaggi non letti per il ruolo corrente (badge bottom-nav).
    Per il cliente ritorna anche `closed_unseen`: numero di ticket appena
    chiusi dall'admin e non ancora "visti" dal cliente (pallino su filtro Chiusi).
    """
    tickets = await store.find("tickets")
    if user.get("role") == "admin":
        n = sum(t.get("unread_admin", 0) for t in tickets if t.get("status") == "open")
        # Per l'admin il "closed_unseen" non ha senso: lo restituisco a 0
        return {"count": n, "closed_unseen": 0}
    n = sum(t.get("unread_customer", 0) for t in tickets
            if t.get("sold_to") == user.get("sold_to_code")
            and t.get("status") == "open")
    closed_unseen = sum(1 for t in tickets
                         if t.get("sold_to") == user.get("sold_to_code")
                         and t.get("status") == "closed"
                         and t.get("closed_unseen", False))
    return {"count": n, "closed_unseen": closed_unseen}


@app.post("/tickets/mark_closed_seen")
async def mark_closed_seen(user: dict = Depends(get_current_user)):
    """Il cliente segna "visto" tutti i suoi ticket chiusi (pallino del chip Chiusi).
    Idempotente. Non-op per admin (closed_unseen non lo riguarda)."""
    if user.get("role") == "admin":
        return {"ok": True, "marked": 0}
    sold_to = user.get("sold_to_code")
    if not sold_to:
        return {"ok": True, "marked": 0}
    tickets = await store.find("tickets", {"sold_to": sold_to, "status": "closed"})
    n = 0
    for t in tickets:
        if t.get("closed_unseen"):
            await store.update_one("tickets", {"id": t["id"]}, {"closed_unseen": False})
            n += 1
    return {"ok": True, "marked": n}


@app.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: str, user: dict = Depends(get_current_user)):
    t = await store.find_one("tickets", {"id": ticket_id})
    if not t:
        raise HTTPException(404, "Ticket non trovato")
    if not _ticket_can_access(t, user):
        raise HTTPException(403, "Non autorizzato")
    return t


@app.post("/tickets/{ticket_id}/messages")
async def add_ticket_message(ticket_id: str, req: TicketMessageReq,
                              user: dict = Depends(get_current_user)):
    """Invia un messaggio nella chat del ticket.
    Comportamento speciale: se il ticket è "closed" e arriva un nuovo messaggio
    (sia da cliente che da admin), il ticket viene RIAPERTO automaticamente.
    In questo modo il cliente che vuole rispondere a una chiusura prematura
    può farlo senza bisogno di chiedere all'admin di riaprire.
    """
    t = await store.find_one("tickets", {"id": ticket_id})
    if not t:
        raise HTTPException(404, "Ticket non trovato")
    if not _ticket_can_access(t, user):
        raise HTTPException(403, "Non autorizzato")

    msg_text = (req.message or "").strip()
    if not msg_text:
        raise HTTPException(400, "Messaggio vuoto")
    now = datetime.now(timezone.utc).isoformat()
    msg = {
        "id": str(uuid.uuid4()),
        "author_id": user["id"],
        "author_role": user.get("role", "customer"),
        "author_name": user.get("username") or user.get("email", ""),
        "text": msg_text,
        "created_at": now,
    }
    msgs = list(t.get("messages") or [])
    msgs.append(msg)
    # Aggiorno contatori: il messaggio è non letto per la controparte
    if user.get("role") == "admin":
        unread_customer = (t.get("unread_customer", 0) or 0) + 1
        unread_admin = t.get("unread_admin", 0) or 0
    else:
        unread_admin = (t.get("unread_admin", 0) or 0) + 1
        unread_customer = t.get("unread_customer", 0) or 0
    patch: Dict[str, Any] = {
        "messages": msgs,
        "updated_at": now,
        "last_message_at": now,
        "unread_admin": unread_admin,
        "unread_customer": unread_customer,
    }
    # Auto-riapertura se chiuso
    reopened = False
    if t.get("status") == "closed":
        patch["status"] = "open"
        patch["closed_at"] = None
        patch["closed_unseen"] = False
        reopened = True
    await store.update_one("tickets", {"id": ticket_id}, patch)
    out = {**t, **patch}
    if reopened:
        out["reopened"] = True
    return out


@app.post("/tickets/{ticket_id}/attachments")
async def add_ticket_attachment(
    ticket_id: str,
    message: str = Form(""),
    file: UploadFile = File(...),
    user: dict = Depends(get_admin_user),
):
    """Admin-only: allega un file a un ticket, con messaggio testuale opzionale.
    Il file viene codificato in base64 e salvato nel corpo del messaggio.
    Limite: 5 MB. Formati suggeriti: immagini, PDF, Excel, Word."""
    import base64
    MAX_BYTES = 5 * 1024 * 1024  # 5 MB
    t = await store.find_one("tickets", {"id": ticket_id})
    if not t:
        raise HTTPException(404, "Ticket non trovato")

    content = await file.read()
    if len(content) > MAX_BYTES:
        raise HTTPException(
            400,
            f"File troppo grande ({len(content)//1024} KB). Limite: 5 MB."
        )

    data_b64 = base64.b64encode(content).decode("ascii")
    now = datetime.now(timezone.utc).isoformat()
    msg = {
        "id": str(uuid.uuid4()),
        "author_id": user["id"],
        "author_role": "admin",
        "author_name": user.get("username") or user.get("email", ""),
        "text": (message or "").strip(),
        "attachment": {
            "filename": file.filename or "allegato",
            "mime": file.content_type or "application/octet-stream",
            "data_b64": data_b64,
        },
        "created_at": now,
    }

    msgs = list(t.get("messages") or [])
    msgs.append(msg)
    unread_customer = (t.get("unread_customer", 0) or 0) + 1
    unread_admin = t.get("unread_admin", 0) or 0
    patch: Dict[str, Any] = {
        "messages": msgs,
        "updated_at": now,
        "last_message_at": now,
        "unread_admin": unread_admin,
        "unread_customer": unread_customer,
    }
    reopened = False
    if t.get("status") == "closed":
        patch["status"] = "open"
        patch["closed_at"] = None
        patch["closed_unseen"] = False
        reopened = True

    await store.update_one("tickets", {"id": ticket_id}, patch)
    out = {**t, **patch}
    if reopened:
        out["reopened"] = True
    return out


@app.patch("/tickets/{ticket_id}/status")
async def set_ticket_status(ticket_id: str, req: TicketStatusReq,
                              _: dict = Depends(get_admin_user)):
    """Solo admin può chiudere/riaprire un ticket.
    Quando l'admin chiude, settiamo closed_unseen=True così il cliente vede
    un pallino sul chip "Chiusi" per accorgersi della chiusura. Alla riapertura
    azzeriamo il flag."""
    t = await store.find_one("tickets", {"id": ticket_id})
    if not t:
        raise HTTPException(404, "Ticket non trovato")
    now = datetime.now(timezone.utc).isoformat()
    patch: Dict[str, Any] = {
        "status": req.status,
        "updated_at": now,
    }
    if req.status == "closed":
        patch["closed_at"] = now
        patch["closed_unseen"] = True
    else:
        patch["closed_at"] = None
        patch["closed_unseen"] = False
    await store.update_one("tickets", {"id": ticket_id}, patch)
    return {**t, **patch}


@app.post("/tickets/{ticket_id}/read")
async def mark_ticket_read(ticket_id: str, user: dict = Depends(get_current_user)):
    """Azzera il contatore unread del ruolo corrente per questo ticket.
    Tipicamente chiamato dal frontend quando l'utente apre la chat.
    Se il cliente apre un ticket chiuso, pulisce anche closed_unseen così
    il pallino sul chip "Chiusi" si aggiorna."""
    t = await store.find_one("tickets", {"id": ticket_id})
    if not t:
        raise HTTPException(404, "Ticket non trovato")
    if not _ticket_can_access(t, user):
        raise HTTPException(403, "Non autorizzato")
    patch: Dict[str, Any] = {}
    if user.get("role") == "admin":
        patch["unread_admin"] = 0
    else:
        patch["unread_customer"] = 0
        if t.get("status") == "closed" and t.get("closed_unseen"):
            patch["closed_unseen"] = False
    await store.update_one("tickets", {"id": ticket_id}, patch)
    return {"ok": True, **patch}


@app.delete("/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, _: dict = Depends(get_admin_user)):
    """Cancellazione hard del ticket (admin only). Utile per pulire spam/test."""
    ok = await store.delete_one("tickets", {"id": ticket_id})
    if not ok:
        raise HTTPException(404, "Ticket non trovato")
    return {"ok": True}


# ── SERVE FRONTEND ────────────────────────────────────────────────────────
# Comodo per uso locale: il backend serve anche Logistic_app.html, manifest, sw, icone.
# In produzione probabilmente sposterai il frontend su CDN / Vercel.
# ─── FOTO ARTICOLI ──────────────────────────────────────────────────────────
# Collection `article_images`. Una foto per article_id (es. "KZ9025").
# Per ora le foto sono caricate dall'admin (in pratica solo per COBA) ma il
# matching è per article_id, quindi se l'articolo è in più clienti la foto
# è visibile a tutti. Storage: base64 in Mongo (file ~100 KB, 200 articoli
# = ~20 MB → ben sotto i 512 MB di Atlas free).
#
# La rotta GET è PUBBLICA: i tag <img src=...> non possono inviare Authorization
# header. Le foto di prodotto non sono dati riservati (sono pubblicate dal
# brand stesso sui suoi cataloghi), quindi accettiamo il trade-off.

@app.get("/article_images")
async def list_article_images(_: dict = Depends(get_admin_user)):
    """Lista articoli con foto (metadati, senza il binary)."""
    imgs = await store.find("article_images")
    return [{
        "article_id":  i.get("article_id", ""),
        "filename":    i.get("filename", ""),
        "mime":        i.get("mime", ""),
        "size":        i.get("size", 0),
        "uploaded_at": i.get("uploaded_at", ""),
    } for i in imgs]


@app.post("/article_images")
async def upload_article_image(
    article_id: str = Form(...),
    file: UploadFile = File(...),
    _: dict = Depends(get_admin_user),
):
    """Upload singolo: associa una foto a un article_id. Sostituisce se esiste."""
    import base64
    aid = (article_id or "").strip().upper()
    if not aid:
        raise HTTPException(400, "article_id richiesto")
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "File vuoto")
    if len(raw) > 12 * 1024 * 1024:
        raise HTTPException(413, "File troppo grande (max 12 MB)")
    mime = (file.content_type or "").lower()
    if not mime.startswith("image/"):
        # Indovino dal filename
        ext = (file.filename or "").lower().rsplit(".", 1)[-1] if "." in (file.filename or "") else ""
        mime = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","webp":"image/webp","gif":"image/gif"}.get(ext, "image/jpeg")
    doc = {
        "article_id":  aid,
        "data_b64":    base64.b64encode(raw).decode("ascii"),
        "mime":        mime,
        "filename":    file.filename or f"{aid}.jpg",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "size":        len(raw),
    }
    # Replace-or-insert
    await store.delete_many("article_images", {"article_id": aid})
    await store.insert("article_images", doc)
    return {"ok": True, "article_id": aid, "size": len(raw), "mime": mime}


@app.post("/article_images/batch")
async def upload_article_images_batch(
    files: List[UploadFile] = File(...),
    _: dict = Depends(get_admin_user),
):
    """Upload batch: per ogni file, article_id = nome file senza estensione
    (uppercase). Es. "KZ9025.jpg" → article_id "KZ9025"."""
    import base64, os.path as op
    results: List[dict] = []
    for f in files:
        name = f.filename or ""
        stem = op.splitext(name)[0]
        aid = (stem or "").strip().upper()
        if not aid:
            results.append({"filename": name, "ok": False, "reason": "Nome file vuoto"})
            continue
        try:
            raw = await f.read()
        except Exception as e:
            results.append({"filename": name, "ok": False, "reason": f"read error: {e}"})
            continue
        if not raw:
            results.append({"filename": name, "ok": False, "reason": "File vuoto"})
            continue
        if len(raw) > 12 * 1024 * 1024:
            results.append({"filename": name, "ok": False, "reason": "File > 12 MB"})
            continue
        mime = (f.content_type or "").lower()
        if not mime.startswith("image/"):
            ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""
            mime = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","webp":"image/webp","gif":"image/gif"}.get(ext, "image/jpeg")
        doc = {
            "article_id":  aid,
            "data_b64":    base64.b64encode(raw).decode("ascii"),
            "mime":        mime,
            "filename":    name,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "size":        len(raw),
        }
        await store.delete_many("article_images", {"article_id": aid})
        await store.insert("article_images", doc)
        results.append({"filename": name, "ok": True, "article_id": aid, "size": len(raw)})
    ok_count = sum(1 for r in results if r["ok"])
    return {"ok": True, "uploaded": ok_count, "total": len(files), "results": results}


@app.delete("/article_images/{article_id}")
async def delete_article_image(article_id: str, _: dict = Depends(get_admin_user)):
    aid = (article_id or "").strip().upper()
    n = await store.delete_many("article_images", {"article_id": aid})
    if not n:
        raise HTTPException(404, "Foto non trovata")
    return {"ok": True, "deleted": n}


@app.get("/article_images/{article_id}.{ext}")
async def get_article_image(article_id: str, ext: str):
    """Serve l'immagine come binary. Rotta PUBBLICA (vedi commento sopra).
    L'estensione nel path è solo cosmetica — il MIME viene dal doc."""
    import base64
    aid = (article_id or "").strip().upper()
    img = await store.find_one("article_images", {"article_id": aid})
    if not img:
        raise HTTPException(404, "Foto non trovata")
    raw = base64.b64decode(img.get("data_b64", ""))
    return Response(
        content=raw,
        media_type=img.get("mime", "image/jpeg"),
        headers={"Cache-Control": "public, max-age=3600"},
    )


@app.get("/Logistic_app.html")
async def serve_html():
    return FileResponse(ROOT_DIR / "Logistic_app.html")


@app.get("/manifest.json")
async def serve_manifest():
    return FileResponse(ROOT_DIR / "manifest.json", media_type="application/json")


@app.get("/service-worker.js")
async def serve_sw():
    return FileResponse(ROOT_DIR / "service-worker.js", media_type="application/javascript")


# /icons/* — opzionale, se hai una cartella icons
icons_dir = ROOT_DIR / "icons"
if icons_dir.exists():
    app.mount("/icons", StaticFiles(directory=icons_dir), name="icons")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
